"""Creación y registro de casos.

`ensure_case` es idempotente: si la carpeta existe, solo asegura que estén las
subcarpetas estándar. Nunca borra contenido del usuario.
"""

from __future__ import annotations

import logging
import os
import re
import shutil
import unicodedata
from dataclasses import asdict, dataclass
from datetime import date
from pathlib import Path
from typing import Any, Callable, Iterator

from .config import (
    CARPETA_ID_TO_PATH,
    CASO_SUBDIRS,
    CRM_BUCKET_CONTESTACION,
    CRM_BUCKET_DEMANDA,
    CRM_BUCKET_MONITORIO_DEMANDA,
    CRM_BUCKET_MONITORIO_OPOSICION,
    CRM_BUCKET_OTROS,
    CRM_BUCKET_PRELIMINARES,
    CRM_FALLBACK_PATH,
    CRM_SUBDIR,
    CRM_TREE,
    INFORME_VIABILIDAD_TIPOS,
    SUBDIRS_ALTA_V1,
    caso_path,
    settings,
)
from .intake_utils import sanitize_filename as _sanitize_filename_util
from .utils import now_iso, read_md, write_md

logger = logging.getLogger(__name__)

# Plantillas de viabilidad (paso 7a). Generadas con
# `python -m scripts.render_plantillas all` desde los YAML canónicos en
# `data/_plantillas/`.
_PLANTILLAS_DIR = settings.project_root / "data" / "_plantillas"
# Presupuesto de longitud de ruta para ficheros que abre Office. El sistema de
# ficheros admite más (``LongPathsEnabled=1``) y Python los abre sin problema,
# pero Excel no es long-path aware y se rinde en 260. Los 240 dejan margen para
# el fichero de bloqueo ``~$…`` y para copias tipo «Copia de …».
RUTA_OFFICE_MAX = 240

_INFORME_TEMPLATE = _PLANTILLAS_DIR / "informe_viabilidad.xlsx"
_CUESTIONARIO_TEMPLATE = _PLANTILLAS_DIR / "cuestionario_viabilidad.xlsx"

# Caracteres prohibidos en nombres de fichero en Windows
_FORBIDDEN_FILENAME_CHARS = '/\\:*?"<>|'


@dataclass
class ExpedienteLink:
    """Referencia a un expediente del CRM vinculado a este caso."""
    id: str                              # ID numérico en sudespacho
    element: str                         # "expedientes_judiciales" | "extrajudiciales"
    input_dir: str                       # subcarpeta en 00_Input (ej. "sudespacho_648")


@dataclass
class CaseMeta:
    case_id: str
    titulo: str
    referencia_crm: str | None = None    # "BaRR3 - Roser 39, 2º (W-030LFT) - Art 20 LAU"
    cliente: str | None = None
    contraparte: str | None = None
    jurisdiccion: str = "civil"
    organo: str | None = None
    cuantia: float | None = None
    drive_link: str | None = None
    drive_remote_path: str | None = None
    drive_ev_team_id: str | None = None   # Shared Drive ID de la carpeta E&V (gdrive_ev)
    drive_ev_folder_id: str | None = None  # Folder ID de la carpeta W-XXXXXX
    drive_ev_folder_name: str | None = None  # Nombre de la carpeta W-XXXXXX (cache de la Drive API)
    drive_ev_drive_id: str | None = None     # Shared Drive ID resuelto por la Drive API (cache)
    direccion: str | None = None         # v2: dirección del inmueble (refactor intake v2)
    id_go: str | None = None             # v2: ID GO de Engel & Völkers (p. ej. "BCN-OS-012905")
    tipo_caso: str | None = None         # v2 paso 7: clave de TIPOS_CASO_ALL (NEGATIVA_OFERTA, BAD_DEBT, …)
    ciudad: str | None = None            # subdivisión por ciudades (Fase 2)
    estado: str = "instruccion"          # instruccion | predemanda | demanda | recurso | archivado
    sudespacho_expedientes: list[dict] = None   # lista de ExpedienteLink serializados
    creado_en: str = ""
    actualizado_en: str = ""
    # Biblioteca de casos — lock de checkout/checkin (DISEÑO_V2 §2.3).
    # Retrocompatibles: un _caso.md preexistente sin estos campos se lee con
    # estos defaults (estado_repositorio → "disponible"). La ruta local COMPLETA
    # NUNCA vive aquí (visible para E&V); solo el hostname. La ruta va al
    # _intake_log.jsonl (§2.2 / gobernanza §3).
    estado_repositorio: str = "disponible"      # disponible | prestado | conflicto
    checkout_user: str | None = None
    checkout_timestamp: str | None = None       # ISO 8601 con zona
    checkout_nonce: str | None = None
    checkout_maquina: str | None = None         # hostname, NO ruta
    checkout_notas: str | None = None
    ultimo_checkin_timestamp: str | None = None
    ultimo_checkin_auditlog: str | None = None  # nombre del AUDITLOG en Drive

    def __post_init__(self):
        if self.sudespacho_expedientes is None:
            self.sudespacho_expedientes = []


# ---------------------------------------------------------------------------
# El indice del caso (`00_Input/_caso.md`): crear y ACTUALIZAR sin destruir
# ---------------------------------------------------------------------------
#
# `MEJORAS #146`, diseno rev. 2:
# docs/superpowers/specs/2026-09-05-caso-md-preservar-al-actualizar-design.md.
#
# Hasta el 2026-09-05 `_write_case_index` RECONSTRUIA el fichero entero desde `CaseMeta`
# cada vez que un registrador (`register_expediente`, `register_drive_ev`,
# `cache_drive_folder_info`) lo llamaba sobre un `_caso.md` que ya existia. Con ello se
# perdian, reproducido por sonda: la nota del abogado en el cuerpo, `bucket_override` y
# cualquier clave top-level ajena, las claves de `meta` que el modelo no conoce
# (`proyeccion_local`: un pull convertia la copia prestada en un caso mas del catalogo),
# los wikilinks que `registrar_outputs` inserta bajo `## Navegacion`, y el estado D8 que
# `update_pull_state` guarda solo en la lista top-level `sudespacho_expedientes`.
#
# La guarda va AQUI, en el sumidero, y no en cada registrador: es la leccion de
# `MEJORAS #153` — una guarda en el envoltorio la rodea el siguiente llamador.
#
# El registrador es dueno de EXACTAMENTE tres fragmentos del cuerpo, los que la plantilla
# deriva de los campos que los registradores escriben: la linea de estado, la linea de IDs
# de Drive E&V y la seccion `## Expedientes sudespacho`. Al actualizar se reescriben solo
# esos tres, localizados por su forma; todo lo demas se conserva linea a linea. La rev. 1
# del diseno intentaba decidir si «alguien habia tocado» el cuerpo comparandolo con la
# plantilla, y la R1 adversarial demostro que esa regla CONGELA el cuerpo tras cualquier
# escritor legitimo del frontmatter o del cuerpo, dejando un indice que lista lo
# desvinculado y omite lo vinculado. Aqui no se clasifica nada: no hay que acertar.

#: Las diez claves del frontmatter que son del indice. Todo lo demas en el fichero es de otro.
_CLAVES_PROPIAS_DEL_INDICE: tuple[str, ...] = (
    "case_id", "tipo", "fase", "fecha", "estado", "ciudad", "referencia_crm",
    "sudespacho_expedientes", "drive", "meta",
)
_ENCABEZADO_EXPEDIENTES = "## Expedientes sudespacho"
_ENCABEZADOS_NAVEGACION = ("## Navegación", "## Navegacion")
#: Va dentro de la seccion (c) porque su contrato es el contrario al del resto del cuerpo,
#: y hay que decirlo donde se lee: una edicion a mano AQUI se pierde en la siguiente
#: actualizacion; en el resto del cuerpo, no.
_AVISO_SECCION_GENERADA = "<!-- sección generada por el registrador: no editar a mano -->"
_PREFIJO_LINEA_DRIVE_EV = "- Drive E&V team:"
_PREFIJO_LINEA_REMOTO = "- Remoto rclone:"


def _linea_estado(meta: CaseMeta) -> str:
    return f"Caso `{meta.case_id}` — estado **{meta.estado}**."


def _linea_drive_ev(meta: CaseMeta) -> str | None:
    if meta.drive_ev_team_id or meta.drive_ev_folder_id:
        return f"- Drive E&V team: `{meta.drive_ev_team_id}` / folder: `{meta.drive_ev_folder_id}`"
    return None


def _seccion_expedientes(meta: CaseMeta) -> list[str]:
    """Lineas de la seccion (c); vacia si no hay expedientes.

    TOTAL sobre lo persistido (R1/H-05): `update_pull_state` crea entradas sin `input_dir`,
    y la version anterior hacia `e['input_dir']`, asi que una sola entrada asi convertia en
    fatal cada registrador sobre ese caso.
    """
    entradas = [e for e in (meta.sudespacho_expedientes or []) if isinstance(e, dict)]
    if not entradas:
        return []
    lineas = [_ENCABEZADO_EXPEDIENTES, "", _AVISO_SECCION_GENERADA]
    for e in entradas:
        eid = e.get("id", "?")
        input_dir = e.get("input_dir") or f"sudespacho_{eid}"
        lineas.append(f"- `{e.get('element', '?')}` ID {eid} → `00_Input/{input_dir}/`")
    return lineas


def _cuerpo_del_indice(meta: CaseMeta) -> str:
    """La plantilla del cuerpo, entera. Solo la usa la CREACION."""
    ref_line = f"- Referencia CRM: **{meta.referencia_crm}**\n" if meta.referencia_crm else ""
    drive_ev = _linea_drive_ev(meta)
    drive_ev_line = f"{drive_ev}\n" if drive_ev else ""
    seccion = _seccion_expedientes(meta)
    exp_lines = ("\n" + "\n".join(seccion) + "\n") if seccion else ""
    return (
        f"# {meta.titulo}\n\n"
        f"{_linea_estado(meta)}\n\n"
        f"{ref_line}"
        f"## Partes\n\n"
        f"- Cliente: {meta.cliente or '_(pendiente)_'}\n"
        f"- Contraparte: {meta.contraparte or '_(pendiente)_'}\n\n"
        f"## Sede\n\n"
        f"- Jurisdicción: {meta.jurisdiccion}\n"
        f"- Órgano: {meta.organo or '_(pendiente)_'}\n"
        f"- Cuantía: {meta.cuantia if meta.cuantia is not None else '_(pendiente)_'}\n\n"
        f"## Fuente documental\n\n"
        f"- Drive: {meta.drive_link or '_(sin enlace)_'}\n"
        f"- Remoto rclone: `{meta.drive_remote_path or '_(no configurado)_'}`\n"
        f"{drive_ev_line}"
        f"{exp_lines}\n"
        f"## Navegación\n\n"
        f"- [[scoring]]\n"
        f"- [[viabilidad]]\n"
        f"- [[hechos_atomicos]]\n"
        f"- [[contradicciones]]\n"
        f"- [[demanda]]\n"
    )


def _actualizar_cuerpo(cuerpo: str, meta: CaseMeta) -> str:
    """Reescribe en `cuerpo` SOLO los tres fragmentos del registrador (diseno §3.3).

    (a) la linea de estado, si esta; (b) la linea de IDs de Drive E&V: se sustituye, se
    retira si ya no hay dato, o se inserta tras `- Remoto rclone:` si existe esa linea;
    (c) la seccion `## Expedientes sudespacho`, desde su encabezado hasta el siguiente `## `
    o el final: se sustituye, se retira si ya no hay expedientes, o se inserta antes de
    `## Navegacion` (o al final si no hay). Ninguna otra linea se toca.
    """
    lineas = cuerpo.split("\n")

    # (a) estado
    prefijo_estado = f"Caso `{meta.case_id}` — estado **"
    for i, ln in enumerate(lineas):
        if ln.startswith(prefijo_estado):
            lineas[i] = _linea_estado(meta)
            break

    # (b) IDs de Drive E&V
    nueva_drive = _linea_drive_ev(meta)
    idx = next((i for i, ln in enumerate(lineas) if ln.startswith(_PREFIJO_LINEA_DRIVE_EV)), None)
    if idx is not None:
        if nueva_drive is None:
            del lineas[idx]
        else:
            lineas[idx] = nueva_drive
    elif nueva_drive is not None:
        j = next((i for i, ln in enumerate(lineas) if ln.startswith(_PREFIJO_LINEA_REMOTO)), None)
        if j is not None:
            lineas.insert(j + 1, nueva_drive)

    # (c) seccion de expedientes
    seccion = _seccion_expedientes(meta)
    ini = next((i for i, ln in enumerate(lineas) if ln.strip() == _ENCABEZADO_EXPEDIENTES), None)
    if ini is not None:
        fin = next((i for i in range(ini + 1, len(lineas)) if lineas[i].startswith("## ")),
                   len(lineas))
        lineas[ini:fin] = (seccion + [""]) if seccion else []
    elif seccion:
        nav = next((i for i, ln in enumerate(lineas) if ln.strip() in _ENCABEZADOS_NAVEGACION), None)
        if nav is not None:
            lineas[nav:nav] = seccion + [""]
        else:
            lineas += [""] + seccion
    return "\n".join(lineas)


def _fusionar_expedientes(persistidos: Any, nuevos: Any) -> list:
    """`sudespacho_expedientes` se fusiona POR ENTRADA (diseno §3.2, R1/H-02).

    La lista top-level es el hogar del estado D8 de `update_pull_state` (`last_sync`,
    `doc_ids`, `by_carpeta`, `errors`); el espejo `meta.sudespacho_expedientes` no lo tiene,
    y dos de los tres registradores reconstruian la lista desde el espejo. Se parte de lo
    persistido, en su orden; cada entrada nueva con el mismo `id` se aplica encima; las nuevas
    sin `id` persistido se anaden; las persistidas que la lista nueva no trae se CONSERVAN
    (ningun registrador borra por esta via: `remove_expediente_link` va por
    `_atomic_write_caso_md`).
    """
    salida: list = [dict(e) if isinstance(e, dict) else e
                    for e in (persistidos if isinstance(persistidos, list) else [])]
    posicion = {str(e.get("id")): i for i, e in enumerate(salida)
                if isinstance(e, dict) and e.get("id") is not None}
    for n in (nuevos if isinstance(nuevos, list) else []):
        if isinstance(n, dict) and n.get("id") is not None and str(n["id"]) in posicion:
            i = posicion[str(n["id"])]
            salida[i] = {**salida[i], **n}
        else:
            salida.append(n)
    return salida


def _frontmatter_del_indice(meta: CaseMeta, expedientes: list) -> dict:
    return {
        "case_id": meta.case_id,
        "tipo": "caso_index",
        "fase": "00_Input",
        "fecha": meta.creado_en,
        "estado": meta.estado,
        "ciudad": meta.ciudad,
        "referencia_crm": meta.referencia_crm,
        "sudespacho_expedientes": expedientes,
        "drive": meta.drive_remote_path,
        "meta": asdict(meta),
    }


def _escribir_indice_atomico(index: Path, fm: dict, cuerpo: str) -> Path:
    """Temporal en el mismo directorio + `os.replace`, como `_atomic_write_caso_md`.

    Las DOS ramas (crear y actualizar) escriben asi: `write_md` en sitio truncaba el
    indice si el proceso moria a mitad, y dejar la creacion «como hoy» dejaba ese agujero
    justo en el escenario que motiva la atomicidad (R1/H-03). El temporal esta en
    `MERGE_EXCLUSIONS` y en el carve-out del plugin para que un huerfano no se trate como
    contenido del expediente (R1/H-07).
    """
    index.parent.mkdir(parents=True, exist_ok=True)
    tmp = index.parent / f"._caso.{os.getpid()}.tmp"
    try:
        write_md(tmp, fm, cuerpo)
        os.replace(tmp, index)
    except Exception:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass
        raise
    return index


def _write_case_index(case_dir: Path, meta: CaseMeta) -> Path:
    """Crea el indice del caso o, si ya existe, lo ACTUALIZA conservando lo que no es suyo.

    Ver el bloque de comentario de arriba y el diseno rev. 2. Un `_caso.md` existente al que
    `read_md` no le encuentra frontmatter (truncado a mitad de escritura) se trata como
    creacion: no tiene nada que conservar y tratarlo como cuerpo lo congelaria corrupto.
    """
    index = case_dir / "00_Input" / "_caso.md"
    if index.exists():
        fm_previo, cuerpo_previo = read_md(index)
        if isinstance(fm_previo, dict) and fm_previo:
            return _actualizar_indice(index, fm_previo, cuerpo_previo, meta)
    expedientes = list(meta.sudespacho_expedientes or [])
    return _escribir_indice_atomico(index, _frontmatter_del_indice(meta, expedientes),
                                    _cuerpo_del_indice(meta))


def _actualizar_indice(index: Path, fm: dict, cuerpo: str, meta: CaseMeta) -> Path:
    from dataclasses import replace as _dc_replace

    expedientes = _fusionar_expedientes(fm.get("sudespacho_expedientes"),
                                        meta.sudespacho_expedientes)
    meta_previo = fm.get("meta") if isinstance(fm.get("meta"), dict) else {}
    # Las claves de `CaseMeta` las manda el dataclass (los registradores ya las han
    # coalescido desde lo persistido); las que el modelo no conoce se conservan.
    meta_dict = {**meta_previo, **asdict(meta)}
    meta_dict["sudespacho_expedientes"] = expedientes
    propias = _frontmatter_del_indice(meta, expedientes)
    propias["meta"] = meta_dict
    # `{**fm, **propias}`: lo ajeno (p. ej. `bucket_override`) queda donde estaba.
    fm_nuevo = {**fm, **propias}
    meta_render = _dc_replace(meta, sudespacho_expedientes=expedientes)
    return _escribir_indice_atomico(index, fm_nuevo, _actualizar_cuerpo(cuerpo, meta_render))


def register_expediente(
    case_id: str,
    expediente_id: str,
    element: str,
) -> str:
    """Registra un expediente del CRM en el índice del caso.

    Añade la entrada a `sudespacho_expedientes` en `_caso.md` si no existe.
    Devuelve el nombre de la subcarpeta de ingesta: `sudespacho_{expediente_id}`.
    Es idempotente: si el expediente ya está registrado, no hace nada.
    """
    input_dir_name = f"sudespacho_{expediente_id}"
    from core.casos.case_locator import buscar
    base = buscar(case_id)
    if base is None:
        return input_dir_name  # el caso no existe — se registrará al crearlo
    index = base / "00_Input" / "_caso.md"

    if not index.exists():
        return input_dir_name  # ensure_case no se llamó aún — se registrará al crearlo

    import yaml as _yaml
    # Un solo parser para el indice (R1/H-03 del diseno de MEJORAS #146): `read_md`, el
    # mismo que usa el sumidero. Con `text.split("---", 2)` un fichero truncado reventaba
    # AQUI con ValueError antes de llegar al sumidero que sabe reconstruirlo.
    fm, _ = read_md(index)
    if not isinstance(fm, dict):
        fm = {}

    expedientes = fm.get("sudespacho_expedientes") or []
    ids_existentes = {str(e.get("id")) for e in expedientes if isinstance(e, dict)}

    if str(expediente_id) not in ids_existentes:
        expedientes.append({
            "id": str(expediente_id),
            "element": element,
            "input_dir": input_dir_name,
        })
        # Re-escribir el índice con el nuevo expediente. Se usa el patrón
        # "known fields" (como register_drive_ev) en lugar de kwargs explícitos:
        # así se preservan TODOS los campos de CaseMeta —incluidos los del lock
        # de checkout (estado_repositorio, checkout_*)— y no se resetean al
        # reescribir el índice de un caso que pudiera estar prestado.
        from dataclasses import fields as _dc_fields

        meta_dict = (fm.get("meta") or {})
        meta_dict["sudespacho_expedientes"] = expedientes
        known = {f.name for f in _dc_fields(CaseMeta)}
        kwargs = {k: v for k, v in meta_dict.items() if k in known}
        kwargs["case_id"] = case_id
        kwargs.setdefault("titulo", case_id)
        kwargs["sudespacho_expedientes"] = expedientes
        kwargs["actualizado_en"] = now_iso()
        meta = CaseMeta(**kwargs)
        _write_case_index(caso_path(case_id), meta)

    return input_dir_name


def _contenido_en(destino: Path, raiz: Path) -> bool:
    """¿`destino` cae dentro de `raiz`? Comparación LÉXICA por componentes, sin tocar disco.

    **Por qué no reutiliza `case_mutex._bajo`, que hace lo mismo.** Era la primera versión, y
    la R2 adversarial demostró que `_bajo` **rechaza a los hijos de una raíz anclada**: hace
    `c.startswith(r + os.sep)`, así que con `r` terminando ya en separador —`C:\\` o un
    recurso UNC— exige dos separadores seguidos y devuelve `False` para todo descendiente
    legítimo (R2/H-02). Medido:

        _bajo(Path('C:/CASOS/EV'),             Path('C:/'))             -> False
        _bajo(Path('//server/share/CASOS/EV'), Path('//server/share/')) -> False

    Con `os.path.commonpath` no hay ese caso especial, y su `ValueError` entre unidades
    distintas **es** la respuesta correcta: no contenido. No se arregla `_bajo` porque lo
    consume el mutex del caso, cuyo radio de daño es otro y no se toca de noche: `MEJORAS #159`.
    """
    d = os.path.normcase(os.path.abspath(str(destino)))
    r = os.path.normcase(os.path.abspath(str(raiz)))
    try:
        return os.path.commonpath([d, r]) == r
    except ValueError:
        return False


def ensure_case(
    case_id: str,
    *,
    titulo: str | None = None,
    referencia_crm: str | None = None,
    cliente: str | None = None,
    contraparte: str | None = None,
    drive_link: str | None = None,
    drive_remote_path: str | None = None,
    cuantia: float | None = None,
    organo: str | None = None,
    direccion: str | None = None,
    id_go: str | None = None,
    tipo_caso: str | None = None,
    ciudad: str | None = None,
    modo: str = "libre",
    raiz_mutex: "Path | None" = None,
) -> Path:
    """Crea (o asegura) la estructura de un caso. Devuelve la ruta del caso.

    Idempotente. Nunca sobrescribe contenido del usuario.

    Refactor intake v2 — paso 7a (informe renombrado en sesión 7, 2026-05-11):
    - Crea solo la base ``00_Input/05_CRM/`` (D7 — andamiaje *lazy* tras la
      reorg 2026-06-10; los buckets se materializan al escribir). Antes creaba
      todas las ramas de ``CRM_TREE`` en eager (D1, derogado).
    - Copia ``data/_plantillas/informe_viabilidad.xlsx`` a
      ``02_Analisis/<nombre>`` salvo que ya haya un informe ahí (con cualquier
      nombre histórico — ver ``_find_informe_existente``). El ``<nombre>`` lo
      decide ``_compose_informe_filename``: ``Informe viabilidad - <id_go>.xlsx``
      si el case_id sigue el formato CRM nuevo y hay ID GO,
      ``_informe_viabilidad.xlsx`` como fallback. El nombre se acortó el
      2026-07-28 (antes llevaba el case_id completo) porque la ruta pasaba de
      260 y Excel no abría el fichero.
    - Si ``tipo_caso ∈ INFORME_VIABILIDAD_TIPOS`` copia además
      ``data/_plantillas/cuestionario_viabilidad.xlsx`` a
      ``02_Analisis/_cuestionario_viabilidad.xlsx``.
    - Pre-rellena REF (``<equipo> - <direccion> (<id_go>)``) y FECHA en el
      informe SOLO cuando se acaba de copiar — preserva trabajo del abogado
      en llamadas posteriores. REF se rellena solo si los tres componentes
      están disponibles (D-7a-2).
    - ``tipo_caso``, ``direccion`` e ``id_go`` se persisten en ``_caso.md``.
      Si el caso ya existe y el kwarg difiere del frontmatter, se actualiza
      vía ``_atomic_write_caso_md`` (D-7a-4).

    Args:
        case_id: Identificador del caso.
        tipo_caso: Clave de ``TIPOS_CASO_ALL``. Gobierna la copia condicional
            del cuestionario de viabilidad.
        direccion, id_go: Persisten en ``_caso.md.meta`` y se usan para
            componer REF del informe.
        (resto): metadatos del caso, opcionales.
    """
    # --- Puerta de alta de V1 (Plan 3A, Task 3). Va ANTES de crear nada.
    #
    # Las tres filas de clase estructura del §25 (#1, #2, #3) no pueden pasar por
    # `core.casos.escritura`: esto crea la raiz ANTES de que exista un caso que el
    # catalogo pueda resolver o un `_caso.md` que el guard pueda leer. De ahi la puerta
    # propia, y de ahi que viva DENTRO de `ensure_case`: su docstring la declara «la UNICA
    # puerta de alta del sistema» y una segunda funcion de alta seria justo lo que esa
    # frase existe para impedir.
    if modo not in ("v1", "libre"):
        raise ValueError(f"modo {modo!r} desconocido; los del §24 D3 son ('v1', 'libre')")
    es_v1 = modo == "v1"
    if es_v1:
        from core.casos import mutex_sesion
        from core.casos.workspace_model import (CaseRef, EscrituraSinMutex,
                                                IdentidadNoUtilizable)

        # El `id_go` EXPLICITO es la identidad, y no se deriva del nombre de la carpeta.
        # En el alta no hay `meta.id_go` todavia, asi que este kwarg es la unica fuente
        # canonica; extraerlo del nombre seria el critico de R14/H14-01 cometido en el
        # unico punto donde no hay metadato que lo desmienta.
        if not id_go:
            raise IdentidadNoUtilizable(
                detalle="el alta en modo v1 exige id_go explicito: sin identidad canonica "
                        "no hay mutex que tomar, y el nombre de la carpeta no es identidad")
        if mutex_sesion.vigente(CaseRef(w_code=id_go), raiz=raiz_mutex) is None:
            raise EscrituraSinMutex(
                w_code=id_go,
                detalle="el alta en modo v1 va bajo el mutex del caso; adquierelo en el "
                        "entrypoint con mutex_sesion.sostenido antes de llamar")

        # --- Concordancia de identidad, ANTES de crear o tocar nada (R15/H15-01).
        #
        # Tener `id_go` y su mutex NO basta: hay que tener el mutex **de este expediente**.
        # Medido: con `meta.id_go = W-OLD01` persistido, el alta aceptaba `id_go=W-NEW01`,
        # **reescribia el metadato** y devolvia el mismo directorio — asi que un proceso con
        # el lock nuevo y otro con el viejo operaban la misma carpeta, los dos creyendose
        # protegidos. Es la falsa exclusion que `IdentidadDiscordante` existe para impedir.
        #
        # Y es la TERCERA vez que aparece la misma propiedad: R14 la cerro en la costura
        # (frontera C0) y yo remedie **ese sitio** en vez de la propiedad — «todo camino que
        # fije identidad comprueba concordancia». El alta es un camino que fija identidad.
        from core.casos import case_locator as _cl
        from core.casos.workspace_model import IdentidadDiscordante

        _canon = CaseRef.normalizar(id_go)
        _existente = _cl.buscar(case_id)
        _declaradas = {_canon}
        if _existente is not None:
            _persistido = str(_cl.read_case_meta(_existente).get("id_go") or "").strip().upper()
            if _persistido:
                _declaradas.add(_persistido)
        _del_nombre = _cl._w_code_de(case_id)
        if _del_nombre:
            _declaradas.add(_del_nombre.strip().upper())
        if len(_declaradas) > 1:
            raise IdentidadDiscordante(
                w_code=_canon,
                detalle="el id_go recibido, el `meta.id_go` persistido y/o el W-code del "
                        "nombre de la carpeta no coinciden; en v1 la identidad de un caso "
                        "NO es un campo que el alta actualice")

    # La UNICA puerta de alta del sistema, y por eso es explicita en el nombre
    # (Task 6 / R7-H7-01). `destino_de_alta` admite que el caso no exista —es su
    # caso normal— y devuelve SU ubicacion si ya existe, lo que impide fabricar
    # una carpeta sombra plana junto a un caso que ya vive en su ciudad.
    # ------------------------------------------------------------------
    # Validacion EN EL SUMIDERO (`MEJORAS #153`/`#154`, diseno rev. 2:
    # docs/superpowers/specs/2026-09-05-validar-en-el-sumidero-design.md).
    #
    # Va aqui y no en los compositores porque el defecto se repitio cuatro veces con la
    # misma forma —«la guarda esta en el envoltorio y el otro llamador la rodea»—: la
    # UI compone su propio `case_id` sin pasar por `componer_case_id`, y el CLI valida la
    # ciudad mientras el core no.
    #
    # ALCANCE, y aqui estaba escrito de mas hasta la R2: esto es el sumidero **del alta
    # nominal**, no del arbol de casos. La primera version de este comentario decia que
    # validando aqui «queda cubierta tambien la puerta que nadie ha escrito todavia», y la
    # R1 lo desmintio con `move_to_city`, que materializa un caso en otra ciudad sin pasar
    # por aqui; la R2 anadio una cuarta puerta (`intake_manual`, que confia en el `lote`
    # que le pasan: `MEJORAS #158`). Las demas estan enumeradas en `#155`/`#156`/`#157` y
    # en el §2 del diseno. Un enunciado mas ancho que lo que la funcion puede prometer es
    # peor que ninguno: invita a no escribir la guarda que falta.
    #
    # ORDEN IMPORTANTE: el `case_id` se valida ANTES de `destino_de_alta`, porque
    # `buscar("")` devuelve la propia raiz y con eso el vacio pasaba entera la validacion.
    # ------------------------------------------------------------------
    from core.utils import exigir_componente_de_ruta
    exigir_componente_de_ruta(case_id, campo="El case_id")
    if ciudad:
        # La ciudad tambien compone el destino, y el contrato no es «que sea un
        # componente» sino **que el caso resultante sea LOCALIZABLE**: con
        # `ciudad="Barcelona/subcarpeta"` el destino cae dentro de la raiz y aun asi
        # `buscar` no lo encuentra, y un segundo alta crea un duplicado (R1/H-03).
        # Se valida contra `_CITY_NAMES`, que es el MISMO conjunto que recorre `buscar`:
        # crear otra lista aqui las haria divergir, que es la enfermedad que esto cura.
        from core.casos.case_locator import _CITY_NAMES
        exigir_componente_de_ruta(ciudad, campo="La ciudad")
        if ciudad not in _CITY_NAMES:
            raise ValueError(
                f"La ciudad {ciudad!r} no esta en el catalogo del localizador, asi que el "
                f"caso quedaria invisible para `buscar`. Admitidas: "
                f"{', '.join(sorted(_CITY_NAMES))}.")

    from core.casos.case_locator import destino_de_alta
    case_dir = destino_de_alta(case_id)
    if not case_dir.exists() and ciudad:
        from core.casos.case_locator import path_for_ciudad
        case_dir = path_for_ciudad(case_id, ciudad)

    # Contencion, y hacen falta LAS DOS mitades. La rev. 2 del diseno puso solo la lexica
    # —reutilizar la primitiva que el repo ya tenia— y el mutante de la junction lo
    # DESMINTIO: `raiz/Enlace` es lexicamente hija de `raiz` aunque los bytes caigan fuera.
    # Reusar algo llamado «contencion» sin comprobar que contenga es el mismo error que
    # este cambio viene a arreglar, cometido un nivel mas arriba.
    #
    #   - LEXICA (`_contenido_en`, aqui al lado): compara por componentes —`CASOS_x` no
    #     esta bajo `CASOS`— y no toca disco, asi que vale con el destino sin crear. Caza
    #     `..` y las rutas absolutas. **NO es `case_mutex._bajo`**, que era la primera
    #     version: la R2 midio que `_bajo` rechaza a los hijos de una raiz anclada
    #     (`MEJORAS #159`), y su docstring lo explica.
    #   - FISICA: resuelve el ancestro EXISTENTE mas cercano y comprueba que siga bajo la
    #     raiz resuelta. Caza los enlaces/junctions, que la lexica no puede ver. Se resuelve
    #     el ancestro y no la hoja porque la hoja normalmente todavia no existe. Se resuelven
    #     los DOS lados: la propia raiz puede ser un enlace (`G:` es Drive Stream).
    #
    # LIMITE DECLARADO: esto contiene el caso y sus ANCESTROS. NO protege de una junction
    # preexistente en un HIJO del caso (p. ej. `00_Input` apuntando fuera), demostrado en la
    # R1 con una junction real de Windows: ahi el `_caso.md` acaba fuera aunque el directorio
    # del caso este contenido. Eso es `MEJORAS #157` y queda FUERA de esta pieza. Decirlo
    # aqui es preferible a que alguien lea contencion total donde no la hay.
    from core.casos.case_locator import _root
    raiz = _root()
    if not _contenido_en(case_dir, raiz):
        raise ValueError(
            f"El destino del caso cae FUERA de la raiz de casos: {case_dir} no esta bajo "
            f"{raiz}. No se deposita un expediente fuera del arbol gobernado.")
    # La comprobacion FISICA solo tiene sentido si hay algo creado que pueda ser un enlace, y
    # el paseo se DETIENE EN LA RAIZ. La primera version subia al primer ancestro existente
    # sin ese tope, asi que con `CASOS_ROOT` todavia sin crear —el primer alta de una maquina
    # nueva— comparaba el PADRE de la raiz contra la raiz y acusaba de escapar por un enlace
    # que no existia (R2/H-01). Un alta que crea su propia raiz es legitima.
    if raiz.exists():
        ancestro = case_dir
        while not ancestro.exists() and _contenido_en(ancestro.parent, raiz):
            ancestro = ancestro.parent
        if ancestro.exists() and not _contenido_en(ancestro.resolve(), raiz.resolve()):
            raise ValueError(
                f"El destino del caso sale de la raiz por un enlace: {case_dir} resuelve a "
                f"{ancestro.resolve()}, fuera de {raiz.resolve()}. No se deposita un "
                "expediente fuera del arbol gobernado.")

    case_dir.mkdir(parents=True, exist_ok=True)

    # Subcarpetas estándar (nivel 1). En V1, solo las que tienen productor DENTRO de la
    # primera vertical (§25 fila #1): las otras seis son andamiaje que luego nadie retira.
    # `90_Notas personales` se conserva eager y es **exencion declarada** —ningun camino de
    # V1 lee ni escribe su contenido, y `core/config.py` ya lo documenta como deliberado—.
    subdirs = SUBDIRS_ALTA_V1 if es_v1 else CASO_SUBDIRS
    for sub in subdirs:
        (case_dir / sub).mkdir(exist_ok=True)

    # Subestructura de 01_Procesado (sala de lectura + MD + cuarentena). Fuera de V1
    # (§25 fila #2): sus productores son `core/sala_lectura.py`, DEPRECADO, y
    # `core/markdown_generator.py`, del pipeline legacy. Ninguno corre en la vertical.
    if not es_v1:
        for sub01 in ("Sala lectura", "MD", "_revisar"):
            (case_dir / "01_Procesado" / sub01).mkdir(exist_ok=True)

    # (ELIMINADO — spec §8 «Scaffolding de ensure_case»: los cajones de entrega y
    # sus roles ya no se crean al alta; los lotes nacen con cada intake y
    # 01_Drive EV lo crea el pull. La base 05_CRM sigue eager (D7), justo abajo.)

    # Base 05_CRM (D7 — andamiaje lazy; los buckets se crean al escribir)
    _ensure_crm_tree_dirs(case_dir)

    index_path = case_dir / "00_Input" / "_caso.md"
    is_new = not index_path.exists()

    if is_new:
        meta = CaseMeta(
            case_id=case_id,
            titulo=titulo or case_id,
            referencia_crm=referencia_crm,
            cliente=cliente,
            contraparte=contraparte,
            drive_link=drive_link,
            drive_remote_path=drive_remote_path,
            cuantia=cuantia,
            organo=organo,
            direccion=direccion,
            id_go=id_go,
            tipo_caso=tipo_caso,
            ciudad=ciudad,
            creado_en=now_iso(),
            actualizado_en=now_iso(),
        )
        _write_case_index(case_dir, meta)
        tipo_caso_eff = tipo_caso
        direccion_eff = direccion
        id_go_eff = id_go
    else:
        # Caso existente: resolver valores efectivos vía coalesce kwarg → persisted.
        try:
            fm, _ = read_md(index_path)
            persisted = (fm.get("meta") or {}) if isinstance(fm, dict) else {}
        except Exception:
            persisted = {}

        tipo_caso_eff = tipo_caso if tipo_caso is not None else persisted.get("tipo_caso")
        direccion_eff = direccion if direccion is not None else persisted.get("direccion")
        id_go_eff = id_go if id_go is not None else persisted.get("id_go")

        # Reescribir frontmatter solo si algún kwarg explícito difiere del persistido (D-7a-4).
        needs_update = (
            (tipo_caso is not None and persisted.get("tipo_caso") != tipo_caso_eff)
            or (direccion is not None and persisted.get("direccion") != direccion_eff)
            or (id_go is not None and persisted.get("id_go") != id_go_eff)
            or (ciudad is not None and persisted.get("ciudad") != ciudad)
        )
        if needs_update:
            def _mutate(fm_in: dict) -> dict:
                meta_in = fm_in.get("meta") or {}
                if tipo_caso is not None:
                    meta_in["tipo_caso"] = tipo_caso_eff
                if direccion is not None:
                    meta_in["direccion"] = direccion_eff
                if id_go is not None:
                    meta_in["id_go"] = id_go_eff
                if ciudad is not None:
                    meta_in["ciudad"] = ciudad
                    fm_in["ciudad"] = ciudad
                fm_in["meta"] = meta_in
                return fm_in
            _atomic_write_caso_md(case_id, _mutate)

    # Fuera de V1 (§25 fila #3): la viabilidad es vertical DIFERIDA (V3), así que copiar
    # sus plantillas al alta deja ficheros de una vertical que todavía no corre. Se
    # devuelve aquí porque lo que queda del cuerpo es plantillas y su pre-rellenado.
    if es_v1:
        return case_dir

    # Copia idempotente de plantillas de viabilidad.
    # Nombre del informe: ``Informe viabilidad - <case_id>.xlsx`` si el
    # case_id sigue el formato CRM nuevo; ``_informe_viabilidad.xlsx`` si
    # no (fallback). Detalle en ``_compose_informe_filename``.
    analisis_dir = case_dir / "02_Analisis"
    informe_existente = _find_informe_existente(analisis_dir)
    if informe_existente is not None:
        # Ya hay informe, quizá con un nombre histórico largo. No se copia otra
        # plantilla (dejaría dos informes, uno en blanco) ni se renombra aquí:
        # el renombrado es del script de migración, que respalda antes de tocar.
        informe_dest = informe_existente
        informe_copiado = False
    else:
        informe_dest = analisis_dir / _compose_informe_filename(case_id, id_go_eff)
        informe_copiado = _copy_plantilla(_INFORME_TEMPLATE, informe_dest)
    _avisar_si_ruta_larga(informe_dest)

    if tipo_caso_eff in INFORME_VIABILIDAD_TIPOS:
        cuestionario_dest = analisis_dir / "_cuestionario_viabilidad.xlsx"
        _copy_plantilla(_CUESTIONARIO_TEMPLATE, cuestionario_dest)

    # Pre-rellenar SOLO si acabamos de copiar el informe — preserva trabajo previo.
    if informe_copiado:
        equipo = _parse_equipo_from_case_id(case_id)
        _prerellenar_informe(
            informe_dest,
            equipo=equipo,
            direccion=direccion_eff,
            id_go=id_go_eff,
        )

    return case_dir


def register_drive_ev(
    case_id: str,
    team_id: str,
    folder_id: str,
) -> None:
    """Registra los IDs del Drive E&V en el frontmatter de _caso.md.

    Almacena drive_ev_team_id y drive_ev_folder_id en el meta del índice.
    Idempotente: si los IDs ya coinciden, no hace nada.
    """
    import yaml as _yaml

    from core.casos.case_locator import buscar
    base = buscar(case_id)
    if base is None:
        return                            # el caso no existe
    index = base / "00_Input" / "_caso.md"
    if not index.exists():
        return                            # el caso existe, `_caso.md` no

    # Un solo parser para el indice (R1/H-03 del diseno de MEJORAS #146): ver register_expediente.
    fm, _ = read_md(index)
    if not isinstance(fm, dict):
        fm = {}

    meta_dict = fm.get("meta") or {}

    # Idempotencia: si ya están los mismos IDs, no reescribir
    if (
        meta_dict.get("drive_ev_team_id") == team_id
        and meta_dict.get("drive_ev_folder_id") == folder_id
    ):
        return

    meta_dict["drive_ev_team_id"] = team_id
    meta_dict["drive_ev_folder_id"] = folder_id

    from dataclasses import fields as _dc_fields

    known = {f.name for f in _dc_fields(CaseMeta)}
    kwargs = {k: v for k, v in meta_dict.items() if k in known}
    kwargs.setdefault("case_id", case_id)
    kwargs.setdefault("titulo", case_id)
    kwargs["actualizado_en"] = now_iso()

    meta = CaseMeta(**kwargs)
    _write_case_index(caso_path(case_id), meta)


def get_drive_ev_ids(case_id: str) -> tuple[str | None, str | None]:
    """Devuelve ``(team_id, folder_id)`` del Drive E&V registrados en ``_caso.md``.

    Lee el frontmatter YAML del índice del caso. Devuelve ``(None, None)``
    si el caso no existe, no tiene índice o aún no tiene carpeta E&V vinculada.

    Args:
        case_id: Identificador del caso.

    Returns:
        Tupla ``(drive_ev_team_id, drive_ev_folder_id)``.
    """
    import yaml as _yaml

    from core.casos.case_locator import buscar
    base = buscar(case_id)
    if base is None:
        return None, None                      # el caso no existe
    index = base / "00_Input" / "_caso.md"
    if not index.exists():
        return None, None                      # el caso existe, `_caso.md` no
    text = index.read_text(encoding="utf-8")
    if not text.startswith("---"):
        return None, None
    try:
        _, fm_raw, _ = text.split("---", 2)
        fm = _yaml.safe_load(fm_raw) or {}
    except Exception:
        return None, None
    meta = fm.get("meta") or {}
    return meta.get("drive_ev_team_id"), meta.get("drive_ev_folder_id")


def get_cached_drive_folder_info(
    case_id: str,
) -> tuple[str | None, str | None]:
    """Devuelve ``(folder_name, drive_id)`` cacheados en ``_caso.md``.

    Returns ``(None, None)`` si no hay cache.
    """
    import yaml as _yaml

    from core.casos.case_locator import buscar
    base = buscar(case_id)
    if base is None:
        return None, None                      # el caso no existe
    index = base / "00_Input" / "_caso.md"
    if not index.exists():
        return None, None                      # el caso existe, `_caso.md` no
    text = index.read_text(encoding="utf-8")
    if not text.startswith("---"):
        return None, None
    try:
        _, fm_raw, _ = text.split("---", 2)
        fm = _yaml.safe_load(fm_raw) or {}
    except Exception:
        return None, None
    meta = fm.get("meta") or {}
    name = meta.get("drive_ev_folder_name")
    drive_id = meta.get("drive_ev_drive_id")
    if name:
        return name, drive_id
    return None, None


def cache_drive_folder_info(
    case_id: str,
    folder_name: str,
    drive_id: str,
) -> None:
    """Persiste nombre y Shared Drive ID de la carpeta E&V en ``_caso.md``.

    Idempotente: si los valores ya coinciden, no reescribe.
    """
    import yaml as _yaml

    from core.casos.case_locator import buscar
    base = buscar(case_id)
    if base is None:
        return                      # el caso no existe
    index = base / "00_Input" / "_caso.md"
    if not index.exists():
        return                      # el caso existe, `_caso.md` no

    # Un solo parser para el indice (R1/H-03 del diseno de MEJORAS #146): ver register_expediente.
    fm, _ = read_md(index)
    if not isinstance(fm, dict):
        fm = {}

    meta_dict = fm.get("meta") or {}

    if (
        meta_dict.get("drive_ev_folder_name") == folder_name
        and meta_dict.get("drive_ev_drive_id") == drive_id
    ):
        return

    meta_dict["drive_ev_folder_name"] = folder_name
    meta_dict["drive_ev_drive_id"] = drive_id

    from dataclasses import fields as _dc_fields

    known = {f.name for f in _dc_fields(CaseMeta)}
    kwargs = {k: v for k, v in meta_dict.items() if k in known}
    kwargs.setdefault("case_id", case_id)
    kwargs.setdefault("titulo", case_id)
    kwargs["actualizado_en"] = now_iso()

    meta = CaseMeta(**kwargs)
    _write_case_index(caso_path(case_id), meta)


def get_case_status(case_id: str) -> dict:
    """Comprueba el estado local del caso.

    Devuelve un dict con:
    - ``local_exists`` (bool): True si la carpeta existe en CASOS_ROOT.
    - ``expedientes`` (list[dict]): expedientes CRM registrados en ``_caso.md``
      (lista vacía si la carpeta no existe o el índice no tiene entradas).

    No lanza excepciones: en caso de error de lectura devuelve estado mínimo.
    """
    import yaml as _yaml

    # Su contrato, escrito en el docstring de arriba, es que NO lanza: es una
    # consulta de estado, y `local_exists` es precisamente la respuesta a «¿existe?».
    # Desde el paso 5 `caso_path` lanza ante un caso ausente, asi que la pregunta
    # se hace con `buscar()`, que devuelve `None` en vez de romper la consulta.
    from core.casos.case_locator import buscar
    case_dir = buscar(case_id)
    local_exists = case_dir is not None
    expedientes: list[dict] = []

    if local_exists:
        index = case_dir / "00_Input" / "_caso.md"
        if index.exists():
            try:
                text = index.read_text(encoding="utf-8")
                if text.startswith("---"):
                    _, fm_raw, _ = text.split("---", 2)
                    fm = _yaml.safe_load(fm_raw) or {}
                    expedientes = [
                        e for e in (fm.get("sudespacho_expedientes") or [])
                        if isinstance(e, dict)
                    ]
            except Exception:
                pass

    return {"local_exists": local_exists, "expedientes": expedientes}


def list_cases() -> list[str]:
    from core.casos.case_locator import list_cases as _list
    return sorted(p.name for p in _list())


# ---------------------------------------------------------------------------
# Biblioteca de casos — lock de checkout/checkin (DISEÑO_V2 §2)
# ---------------------------------------------------------------------------
#
# El "hecho vigente" del lock vive en `_caso.md` (autoridad única). Estas
# funciones operan sobre el `_caso.md` del árbol local (`caso_path`); el frontal
# (CLI/skill) las aplica sobre la copia del Drive tras leerla por API. La
# mutación usa `_atomic_write_caso_md` (preserva TODO el frontmatter y solo toca
# `meta`), nunca `_write_case_index` (que reconstruye y podría descartar campos
# no-CaseMeta como `bucket_override`).
#
# Cero I/O contra Drive aquí tampoco: solo el fichero local. La validación de
# transición, la tabla de estados y la MUTACIÓN del lock (fm→fm) viven en
# `core.repository_checkout` (cerebro puro): estas funciones solo hacen el I/O
# (leer/escribir `_caso.md`) y delegan la lógica. El frontal CLI reutiliza los
# mismos mutadores puros sobre el `_caso.md` del Drive (pull → mutar → push).


def _read_fm(case_id: str) -> dict:
    """Devuelve el frontmatter completo de `_caso.md` (o {} si no hay/no parsea)."""
    from core.casos.case_locator import buscar
    base = buscar(case_id)
    if base is None:
        return {}                      # el caso no existe
    index = base / "00_Input" / "_caso.md"
    if not index.exists():
        return {}                      # el caso existe, `_caso.md` no
    try:
        fm, _ = read_md(index)
    except Exception:
        return {}
    return fm if isinstance(fm, dict) else {}


def leer_estado_repositorio(case_id: str) -> str:
    """Estado del lock del caso. `"disponible"` por defecto (retrocompatible)."""
    from .repository_checkout import estado_de_fm
    return estado_de_fm(_read_fm(case_id))


def leer_lock(case_id: str) -> dict[str, Any]:
    """Devuelve los campos del lock con defaults (nunca lanza).

    Un `_caso.md` sin los campos nuevos se lee como caso disponible sin
    checkout activo.
    """
    from .repository_checkout import leer_lock_de_fm
    return leer_lock_de_fm(_read_fm(case_id))


def escribir_lock(
    case_id: str,
    *,
    user: str,
    timestamp: str,
    nonce: str,
    maquina: str | None = None,
    notas: str | None = None,
) -> dict[str, Any]:
    """Adquiere el lock: transición `disponible → prestado` y escribe los campos.

    Valida la transición contra el estado vigente (rechaza el doble checkout:
    `prestado → prestado` lanza `TransicionInvalida`). Escribe `checkout_user`,
    `checkout_timestamp` (ISO con zona), `checkout_nonce`, y opcionalmente
    `checkout_maquina` (hostname) y `checkout_notas`. NO escribe ruta local
    (§2.2). Devuelve el lock resultante.

    El write-then-verify con nonce (releer el Drive y confirmar el ganador) lo
    orquesta el frontal con `repository_checkout.verificar_nonce`.
    """
    from .config import ESTADO_REPO_PRESTADO
    from .repository_checkout import aplicar_lock_prestado, validar_transicion

    validar_transicion(leer_estado_repositorio(case_id), ESTADO_REPO_PRESTADO)
    _atomic_write_caso_md(case_id, lambda fm: aplicar_lock_prestado(
        fm, user=user, timestamp=timestamp, nonce=nonce, maquina=maquina, notas=notas))
    return leer_lock(case_id)


def liberar_lock(
    case_id: str,
    *,
    timestamp: str,
    auditlog: str | None = None,
) -> dict[str, Any]:
    """Libera el lock: transición `→ disponible` tras un checkin verificado.

    Limpia los campos `checkout_*` y fija `ultimo_checkin_timestamp` /
    `ultimo_checkin_auditlog`. Valida la transición (desde `prestado` o
    `conflicto`).
    """
    from .config import ESTADO_REPO_DISPONIBLE
    from .repository_checkout import aplicar_lock_liberado, validar_transicion

    validar_transicion(leer_estado_repositorio(case_id), ESTADO_REPO_DISPONIBLE)
    _atomic_write_caso_md(case_id, lambda fm: aplicar_lock_liberado(
        fm, timestamp=timestamp, auditlog=auditlog))
    return leer_lock(case_id)


def cancelar_checkout(case_id: str, *, timestamp: str) -> dict[str, Any]:
    """Cancela el checkout descartando el trabajo local (runbook §7.1).

    Transición `→ disponible` SIN registrar checkin (no hubo merge). Limpia el
    lock. El aviso «el trabajo local se descarta» lo da el frontal antes de
    llamar; el evento `checkout_cancelado` lo emite el frontal en el log. El
    parámetro `timestamp` se acepta para simetría de la API y trazabilidad del
    frontal, aunque el estado resultante no lo persiste.
    """
    from .config import ESTADO_REPO_DISPONIBLE
    from .repository_checkout import aplicar_lock_cancelado, validar_transicion

    validar_transicion(leer_estado_repositorio(case_id), ESTADO_REPO_DISPONIBLE)
    _atomic_write_caso_md(case_id, aplicar_lock_cancelado)
    return leer_lock(case_id)


def marcar_conflicto(case_id: str) -> dict[str, Any]:
    """Transición `prestado → conflicto` (checkin con conflictos; local SE CONSERVA)."""
    from .config import ESTADO_REPO_CONFLICTO
    from .repository_checkout import aplicar_estado, validar_transicion

    validar_transicion(leer_estado_repositorio(case_id), ESTADO_REPO_CONFLICTO)
    _atomic_write_caso_md(case_id, lambda fm: aplicar_estado(fm, ESTADO_REPO_CONFLICTO))
    return leer_lock(case_id)


def es_copia_prestada(case_id: str) -> bool:
    """¿La ruta resuelta de este caso es una copia local conocida, y no el canon?

    Lo contesta el **registro privado de workspaces** (Fase 1), que es el SSOT de qué
    copias locales conoce esta máquina. El canon nunca está ahí: el resolver rechaza
    registrar una ruta bajo el catálogo (`WORKSPACE_UNDER_CATALOG_ROOT`).

    ## El discriminante que NO vale, y por qué lo escribo aquí

    El primer intento fue la presencia de **`MANIFEST_CHECKOUT.json`**. Parecía la marca
    inequívoca —la escribe `cmd_checkout`, la lee el checkin como baseline— y **abría un
    agujero de autorización**: `cmd_checkout` sube además una copia del manifiesto **al
    Drive** («debe sobrevivir a la muerte del Desktop», §3.3), así que mientras un caso
    está prestado el fichero está en **las dos copias**. Discriminar por él desactivaba el
    guard sobre el **canon** y justo **mientras otro lo tenía tomado**, que es el único
    momento en que hace falta.

    ## Falla cerrado, y eso es una decisión

    Cualquier fallo al consultar el registro —ilegible, ausente, ruta rara— devuelve
    `False`, o sea «trátalo como el canon» y por tanto **desvía**. «No puedo saberlo» no es
    «no lo es»: es la misma polaridad que el propio registro se aplica (`RegistryUnreadable`
    en vez de `[]`).

    ## Lo que esto NO cubre, declarado

    Un checkout **anterior al registro y sin adoptar** no consta, así que sigue desviando.
    Es deliberado: el sistema no adivina sobre qué copia está. La vía de desbloqueo es
    explícita y existe — `core.casos.workspace_adopcion` (la puerta del §15).
    """
    import os

    try:
        from .casos.case_locator import buscar
        from .casos.workspace_registry import WorkspaceRegistry, raiz_por_defecto

        case_dir = buscar(case_id)
        if case_dir is None:
            return False
        objetivo = os.path.normcase(os.path.abspath(str(case_dir)))
        registro = WorkspaceRegistry(raiz_por_defecto(), ahora="")
        return any(
            os.path.normcase(os.path.abspath(str(e.local_path))) == objetivo
            for e in registro.cargar()
        )
    except Exception:                                   # noqa: BLE001 - falla cerrado
        return False


def guard_escritura(
    case_id: str,
    ruta_relativa: str,
    origen: str,
    *,
    es_protocolo: bool = False,
    emitir_evento: bool = True,
):
    """Guard de escritura del pipeline/UI/intake (DISEÑO_V2 §6).

    Todo punto que escriba en un caso del Drive DEBE llamar aquí antes de
    escribir. Lee el estado_repositorio vigente y decide (vía el cerebro puro
    ``decidir_escritura``) si la escritura procede o se desvía a la bandeja
    ``_pendiente_checkin/<origen>/<ruta>``. Cuando desvía, emite un evento
    ``pendiente_checkin`` en ``_intake_log.jsonl`` (salvo ``emitir_evento=False``).
    El propio PROTOCOLO (lock/log/bandeja) está exento con ``es_protocolo=True``.

    Uso típico::

        d = guard_escritura(case_id, rel, "intake")
        destino = (caso_path(case_id) / d.ruta_bandeja) if d.desviar \
                  else (caso_path(case_id) / rel)

    Returns:
        ``repository_checkout.DecisionEscritura``.
    """
    from .intake_log import append_event
    from .repository_checkout import (DecisionEscritura, decidir_escritura,
                                      evento_pendiente_details)

    if es_copia_prestada(case_id):
        return DecisionEscritura(
            permitido=True, desviar=False, ruta_bandeja=None, evento=None,
            motivo="copia local prestada: la bandeja no aplica (MEJORAS #96)",
        )

    estado = leer_estado_repositorio(case_id)
    decision = decidir_escritura(estado, ruta_relativa, origen, es_protocolo=es_protocolo)
    if decision.desviar and emitir_evento and decision.evento:
        append_event(case_id, decision.evento, details=evento_pendiente_details(
            origen=origen, ruta_bandeja=decision.ruta_bandeja, ruta_original=ruta_relativa))
    return decision


def dir_intake(
    case_id: str,
    rel_base: str,
    origen: str,
    *,
    es_protocolo: bool = False,
) -> Path:
    """Directorio de intake efectivo bajo el caso, aplicando el guard §6.

    Envoltura ergonómica de :func:`guard_escritura` para los puntos de intake
    que escriben en un **directorio base** (WhatsApp, Drive EV, email, CRM): si
    el caso está prestado/conflicto devuelve la ruta dentro de
    ``_pendiente_checkin/<origen>/<rel_base>`` (y registra el desvío); si está
    disponible, devuelve ``caso/<rel_base>``. El caller solo tiene que usar el
    Path devuelto como base de sus escrituras.
    """
    base = caso_path(case_id)
    decision = guard_escritura(case_id, rel_base, origen, es_protocolo=es_protocolo)
    if decision.desviar:
        return base / decision.ruta_bandeja
    return base / rel_base


# ---------------------------------------------------------------------------
# Refactor intake v2 — helpers
# ---------------------------------------------------------------------------
#
# Estrategia y decisiones cerradas: project_intake_estructura_v2.md (memoria)
# + docs/INTEGRACION_SUDESPACHO.md §13.
#
# Casos antiguos (BaRR3, MaRS15) están congelados (D9): el sync v2 debe llamar
# a `is_legacy_intake_v1()` antes de operar y bloquear con mensaje claro.


def _normalize_label(s: str | None) -> str:
    """Normalización tolerante (M4-Q3).

    Lowercase, elimina acentos, strip. Útil para casar `id_carpeta_label`
    devuelto por `/api/element_registries/gdocu` con los nombres del árbol
    local (CRM_TREE) sin sensibilidad a capitalización ni acentos.
    """
    if not s:
        return ""
    nfkd = unicodedata.normalize("NFKD", s)
    ascii_s = nfkd.encode("ascii", "ignore").decode("ascii")
    return ascii_s.strip().lower()


def _walk_crm_tree(
    tree: dict[str, dict],
    prefix: str = "",
) -> Iterator[tuple[str, str]]:
    """Recorre CRM_TREE en DFS devolviendo (nombre_nodo, ruta_completa).

    Las rutas usan separador "/" (D11). Incluye nodos intermedios y hojas.
    """
    for name, children in tree.items():
        full = f"{prefix}/{name}" if prefix else name
        yield (name, full)
        if isinstance(children, dict) and children:
            yield from _walk_crm_tree(children, full)


def _find_branches_by_label(label: str) -> list[str]:
    """Devuelve todas las rutas del árbol cuyo nombre normalizado coincide.

    Recolecta todos los nodos (intermedios + hojas) cuyo nombre normalizado
    es igual al `label` normalizado. La heurística solo aplica si len == 1
    (ambigüedad → fallback).
    """
    target = _normalize_label(label)
    if not target:
        return []
    matches: list[str] = []
    for name, full_path in _walk_crm_tree(CRM_TREE):
        if _normalize_label(name) == target:
            matches.append(full_path)
    return matches


def _bucket_for(rama_canonica: str) -> str:
    """Aplana una rama canónica de ``CRM_TREE`` al bucket procesal plano (D5/D6).

    Routing por **rama completa**, no por etiqueta-hoja: la hoja ``"Demanda"``
    aparece en tres ramas distintas (Declarativo, Monitorio, Preliminares) y
    ``"Oposicion"`` en dos, así que solo la rama completa desambigua.

    ``Preliminares`` está en **lista de exclusión explícita** y se comprueba
    ANTES del match genérico de ``Demanda``/``Oposicion``: la "demanda" de
    diligencias preliminares (solicitud de DP) **nunca** cae en ``01_Demanda``
    → va a ``05_Diligencias_Preliminares``.

    El resto de ramas (``General``, ``Civil``, ``1ª Instancia``, ``Documentos``,
    ``Documentacion RGPD LOPD``, ``Apelacion``, ``Ejecucion``, ``Penal/*``…)
    cae al bucket plano ``99_Otros``.

    Args:
        rama_canonica: ruta con separador ``"/"`` tal como la almacena
            ``CARPETA_ID_TO_PATH`` o la devuelve ``_find_branches_by_label``
            (p. ej. ``"Civil/1ª Instancia/Declarativo/Demanda"``). Tolerante a
            acentos/capitalización vía ``_normalize_label``.

    Returns:
        Nombre del bucket (un solo segmento), p. ej. ``"01_Demanda"``.
    """
    norm = [_normalize_label(p) for p in rama_canonica.split("/") if p]

    # Exclusión explícita: cualquier rama bajo Preliminares → 05 (D6).
    if "preliminares" in norm:
        return CRM_BUCKET_PRELIMINARES

    leaf = norm[-1] if norm else ""
    parent = norm[-2] if len(norm) >= 2 else ""

    if parent == "declarativo" and leaf == "demanda":
        return CRM_BUCKET_DEMANDA
    if parent == "declarativo" and leaf == "oposicion":
        return CRM_BUCKET_CONTESTACION
    if parent == "monitorio" and leaf == "demanda":
        return CRM_BUCKET_MONITORIO_DEMANDA
    if parent == "monitorio" and leaf == "oposicion":
        return CRM_BUCKET_MONITORIO_OPOSICION

    return CRM_BUCKET_OTROS


def resolve_bucket(
    id_carpeta: str | int | None = None,
    id_carpeta_label: str | None = None,
) -> tuple[str | None, str]:
    """Resuelve un documento del CRM a su bucket plano (sin ruta de caso).

    Fuente única de verdad de la resolución carpeta→bucket, compartida por
    ``crm_branch_path`` (que añade la ruta del caso) y por el detector de
    conjunto (D9), que necesita el bucket de un doc sin construir la ruta.

    Args:
        id_carpeta: ID numérico de la carpeta Gdocu (string o int).
        id_carpeta_label: label leaf-only (puede venir vacío).

    Returns:
        Tupla ``(bucket, kind)``: ``bucket`` es el nombre del bucket plano
        (p. ej. ``"01_Demanda"``) o ``None`` si ni el ID ni el label resuelven
        a una rama única; ``kind`` ∈ ``{"id_mapping", "label_heuristic",
        "fallback"}``.
    """
    if id_carpeta is not None:
        key = str(id_carpeta).strip()
        if key in CARPETA_ID_TO_PATH:
            return _bucket_for(CARPETA_ID_TO_PATH[key]), "id_mapping"

    if id_carpeta_label:
        candidates = _find_branches_by_label(id_carpeta_label)
        if len(candidates) == 1:
            return _bucket_for(candidates[0]), "label_heuristic"

    return None, "fallback"


_VALID_BUCKETS: frozenset[str] = frozenset({
    CRM_BUCKET_DEMANDA,
    CRM_BUCKET_CONTESTACION,
    CRM_BUCKET_MONITORIO_DEMANDA,
    CRM_BUCKET_MONITORIO_OPOSICION,
    CRM_BUCKET_PRELIMINARES,
    CRM_BUCKET_OTROS,
})


def read_bucket_overrides(case_id: str) -> dict[str, str]:
    """Lee el override local ``doc_id → bucket`` del frontmatter de ``_caso.md`` (D11).

    El letrado puede forzar el bucket de un documento mal archivado en el CRM
    editando a mano el campo ``bucket_override`` (mapa ``doc_id: bucket``) del
    frontmatter de ``00_Input/_caso.md``. El override se respeta **por encima**
    de la carpeta del CRM y **sin tocar el CRM remoto** — es un parche local.

    Solo se devuelven entradas cuyo bucket sea uno de los buckets válidos
    (``_VALID_BUCKETS``): un valor inválido se descarta en silencio para no
    materializar una carpeta espuria. Las claves se normalizan a ``str``.

    Returns:
        Mapa ``{doc_id: bucket}`` (vacío si no hay caso, índice o campo).
    """
    import yaml as _yaml

    from core.casos.case_locator import buscar
    base = buscar(case_id)
    if base is None:
        return {}                      # el caso no existe
    index = base / "00_Input" / "_caso.md"
    if not index.exists():
        return {}                      # el caso existe, `_caso.md` no
    text = index.read_text(encoding="utf-8")
    if not text.startswith("---"):
        return {}
    try:
        _, fm_raw, _ = text.split("---", 2)
        fm = _yaml.safe_load(fm_raw) or {}
    except Exception:
        return {}
    raw = fm.get("bucket_override")
    if not isinstance(raw, dict):
        return {}
    return {
        str(k).strip(): v
        for k, v in raw.items()
        if isinstance(v, str) and v in _VALID_BUCKETS
    }


def crm_branch_path(
    case_id: str,
    *,
    id_carpeta: str | int | None = None,
    id_carpeta_label: str | None = None,
    expediente_id: str | int | None = None,
    doc_id: str | int | None = None,
    overrides: dict[str, str] | None = None,
) -> tuple[Path, str]:
    """Resuelve la ruta destino dentro de ``00_Input/05_CRM/`` para un documento del CRM.

    Estrategia híbrida (M4 + M5; ver §13.4 de docs/INTEGRACION_SUDESPACHO.md).
    Tras la reorg 2026-06-10 (D5/D6) el destino es un **bucket plano** de un
    nivel, no la rama profunda: se resuelve la rama canónica y se aplana con
    ``_bucket_for``.

    0. **Override local del letrado (D11)**: si ``doc_id`` está en el mapa
       ``bucket_override`` de ``_caso.md`` (o en ``overrides`` si se pasa
       pre-leído), ese bucket manda **por encima** de la carpeta del CRM.
    1. Lookup directo en ``CARPETA_ID_TO_PATH`` por ``id_carpeta`` → rama →
       ``_bucket_for`` → bucket.
    2. Heurística por ``id_carpeta_label`` si la coincidencia en ``CRM_TREE``
       es única (ambigüedad → fallback) → rama → ``_bucket_for`` → bucket.
    3. Fallback ``05_CRM/99_Sin categoria/<expediente_id>/``. El caller debe
       escribir un evento ``category_unknown`` en ``_intake_log.jsonl`` (M10)
       cuando ``kind == "fallback"`` para descubrimiento progresivo de IDs.

    Args:
        case_id: ID del caso.
        id_carpeta: ID numérico devuelto por ``/api/element_registries/gdocu``
            (acepta string o int — se normaliza con ``str().strip()``).
        id_carpeta_label: label leaf-only del mismo endpoint (puede venir vacío).
        expediente_id: ID del expediente CRM, usado solo para el fallback.
        doc_id: ID del documento, para consultar el override local (D11).
        overrides: mapa ``doc_id → bucket`` ya leído. Si es ``None`` y se pasa
            ``doc_id``, se lee de ``_caso.md`` con ``read_bucket_overrides``;
            el caller de un bucle (pull) debe leerlo una vez y pasarlo para
            evitar I/O por-documento.

    Returns:
        Tupla ``(path, kind)`` donde ``kind`` ∈ ``{"override", "id_mapping",
        "label_heuristic", "fallback"}``. El path siempre está bajo
        ``<casos_root>/<case_id>/00_Input/05_CRM/`` y apunta a un bucket plano
        (o al fallback ``99_Sin categoria/<exp>``).
    """
    base = caso_path(case_id) / "00_Input" / CRM_SUBDIR

    # 0. Override local del letrado (D11) — por encima de la carpeta del CRM.
    if doc_id is not None:
        ov_map = overrides if overrides is not None else read_bucket_overrides(case_id)
        ov_bucket = ov_map.get(str(doc_id).strip())
        if ov_bucket in _VALID_BUCKETS:
            return base / ov_bucket, "override"

    # 1+2. Resolución carpeta→bucket (id_mapping / label_heuristic) — fuente
    # única en resolve_bucket.
    bucket, kind = resolve_bucket(id_carpeta, id_carpeta_label)
    if bucket is not None:
        return base / bucket, kind

    # 3. Fallback
    fallback = base / CRM_FALLBACK_PATH
    if expediente_id is not None:
        fallback = fallback / str(expediente_id)
    return fallback, "fallback"


def is_legacy_intake_v1(case_id: str) -> bool:
    """Detecta si el caso tiene estructura v1 de intake CRM (D9).

    Devuelve True si existe alguna subcarpeta ``00_Input/sudespacho_*/``.
    Casos v1 están congelados — el pull v2 debe bloquearse con mensaje claro
    en UI (BaRR3, MaRS15). Migración manual: borrar ``sudespacho_*/`` +
    ``force-pull`` v2.
    """
    from core.casos.case_locator import buscar
    base = buscar(case_id)
    if base is None:
        return False                      # el caso no existe
    input_dir = base / "00_Input"
    if not input_dir.exists():
        return False                      # el caso existe, `00_Input` no
    for child in input_dir.iterdir():
        if child.is_dir() and child.name.startswith("sudespacho_"):
            return True
    return False


def _atomic_write_caso_md(
    case_id: str,
    mutator: Callable[[dict], dict | None],
) -> Path:
    """Escritura atómica del frontmatter de _caso.md (D10).

    Lee ``fm + body``, aplica ``mutator(fm)`` y escribe a un archivo temporal
    dentro del mismo directorio; cierra la operación con ``os.replace``. Sin
    lock, sin versionado. La idempotencia es responsabilidad del mutator.

    Args:
        case_id: ID del caso. ``_caso.md`` debe existir.
        mutator: callable que recibe el frontmatter parseado (dict). Puede
            mutar in-place o devolver un dict nuevo; si devuelve ``None`` se
            asume mutación in-place.

    Returns:
        Path al archivo escrito.

    Raises:
        FileNotFoundError: si ``_caso.md`` no existe.
    """
    # `localizar` lanza el error estructurado del §10 si el caso no existe. El
    # mensaje que habia aqui interpolaba el `case_id` ENTERO, y un `case_id` es
    # `BaXXX - <direccion> - (W-XXXXX) - <tipo>`: publicaba la direccion del
    # inmueble, que el §16 prohibe.
    from core.casos.case_locator import localizar
    index = localizar(case_id) / "00_Input" / "_caso.md"
    if not index.exists():
        raise FileNotFoundError("falta `_caso.md` en el caso")

    fm, body = read_md(index)
    new_fm = mutator(fm)
    if new_fm is not None:
        fm = new_fm

    # Stamp de actualización en meta — consistente con register_expediente y
    # register_drive_ev. No tocamos meta si el mutator no la creó.
    if isinstance(fm.get("meta"), dict):
        fm["meta"]["actualizado_en"] = now_iso()

    # Temp en el mismo directorio para que os.replace sea atómico (Windows + POSIX).
    # PID en el nombre evita colisiones si hubiera procesos concurrentes.
    tmp_path = index.parent / f"._caso.{os.getpid()}.tmp"
    try:
        write_md(tmp_path, fm, body)
        os.replace(tmp_path, index)
    except Exception:
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass
        raise
    return index


def _find_expediente_entry(
    expedientes: list[Any],
    expediente_id: str | int,
) -> tuple[int, dict | None]:
    """Localiza un entry de expediente por ID dentro de la lista del frontmatter.

    Devuelve ``(idx, entry)`` o ``(-1, None)`` si no existe.
    """
    target = str(expediente_id)
    for idx, e in enumerate(expedientes):
        if isinstance(e, dict) and str(e.get("id")) == target:
            return idx, e
    return -1, None


def read_pull_state(
    case_id: str,
    expediente_id: str | int,
) -> dict | None:
    """Devuelve el estado del pull para un expediente concreto (M3 / D8).

    Lee ``sudespacho_expedientes`` del frontmatter de ``_caso.md`` y busca el
    entry cuyo ``id`` coincide con ``expediente_id``. Devuelve ``None`` si el
    caso no existe, no tiene índice, o el expediente no está vinculado.

    Schema D8 del entry: ``{id, element, linked_at, last_sync,
    documents_total_crm, doc_ids, by_carpeta, errors}``.
    """
    from core.casos.case_locator import buscar
    base = buscar(case_id)
    if base is None:
        return None                      # el caso no existe
    index = base / "00_Input" / "_caso.md"
    if not index.exists():
        return None                      # el caso existe, `_caso.md` no
    try:
        fm, _ = read_md(index)
    except Exception:
        return None
    expedientes = fm.get("sudespacho_expedientes") or []
    _, entry = _find_expediente_entry(expedientes, expediente_id)
    return entry


def update_pull_state(
    case_id: str,
    expediente_id: str | int,
    *,
    element: str | None = None,
    last_sync: str | None = None,
    documents_total_crm: int | None = None,
    doc_ids: list[Any] | None = None,
    by_carpeta: dict[str, Any] | None = None,
    errors: list[Any] | None = None,
) -> dict:
    """Crea o actualiza el entry de pull state para un expediente (M3 / D8).

    Schema D8 del entry: ``{id, element, linked_at, last_sync,
    documents_total_crm, doc_ids, by_carpeta, errors}``.

    Semántica:
    - ``linked_at`` se fija en la primera escritura y no se sobrescribe.
    - Cualquier kwarg con valor ``None`` se ignora (campo se conserva).
    - ``errors`` sobrescribe (D12: el state es estado actual; el histórico
      forense vive en ``_intake_log.jsonl`` vía M10).
    - ``element`` es obligatorio al crear el entry. Si se proporciona en una
      actualización, sobrescribe el valor existente.

    Args:
        case_id: ID del caso. ``_caso.md`` debe existir.
        expediente_id: ID numérico del expediente CRM (string o int).
        element: ``"expedientes_judiciales"`` | ``"extrajudiciales"``.
        last_sync: timestamp ISO-8601 del último pull.
        documents_total_crm: total de documentos vistos en el CRM
            (puede superar a ``len(doc_ids)`` por dedup M9).
        doc_ids: IDs numéricos de los documentos descargados.
        by_carpeta: dict ``{ruta_local_canónica: count}`` (D11). Los counts
            incluyen aliases del manifest M9.
        errors: lista de errores del último pull. Sobrescribe el campo.

    Returns:
        El entry final tras la actualización.

    Raises:
        FileNotFoundError: si ``_caso.md`` no existe.
        ValueError: si el entry es nuevo y no se proporciona ``element``.
    """
    target_id = str(expediente_id)
    captured: dict[str, Any] = {}

    def _mutate(fm: dict) -> dict:
        expedientes = fm.get("sudespacho_expedientes")
        if not isinstance(expedientes, list):
            expedientes = []
            fm["sudespacho_expedientes"] = expedientes

        idx, entry = _find_expediente_entry(expedientes, target_id)
        is_new = entry is None
        if entry is None:
            entry = {"id": target_id}

        if is_new:
            if not element:
                raise ValueError(
                    f"element requerido al vincular un expediente nuevo "
                    f"(case={case_id!r}, expediente={target_id!r})"
                )
            entry["element"] = element
            entry["linked_at"] = now_iso()
            entry.setdefault("doc_ids", [])
            entry.setdefault("by_carpeta", {})
            entry.setdefault("errors", [])
        elif element is not None and entry.get("element") != element:
            entry["element"] = element

        if last_sync is not None:
            entry["last_sync"] = last_sync
        if documents_total_crm is not None:
            entry["documents_total_crm"] = documents_total_crm
        if doc_ids is not None:
            entry["doc_ids"] = list(doc_ids)
        if by_carpeta is not None:
            entry["by_carpeta"] = dict(by_carpeta)
        if errors is not None:
            entry["errors"] = list(errors)

        if is_new:
            expedientes.append(entry)
        else:
            expedientes[idx] = entry

        captured["entry"] = entry
        return fm

    _atomic_write_caso_md(case_id, _mutate)
    return captured["entry"]


# ---------------------------------------------------------------------------
# Refactor intake v2 — paso 7a: scaffolding del caso (CRM_TREE + plantillas)
# ---------------------------------------------------------------------------
#
# Helpers privados que dan soporte a `ensure_case` v2. Decisiones D-7a-1 a
# D-7a-7 cerradas con el usuario en sesión 2026-05-11.


# Coincide con el prefijo de equipo del case_id nuevo:
# `BaRR3 - Dirección (W-XXXXXX) - Tipo` → "BaRR3".
_EQUIPO_RE = re.compile(r"^[A-Z][a-zA-Z][A-Z]{2}\d+$")

# ID GO del CRM: ``W-`` + 6 alfanuméricos en mayúsculas (W-02XOR7, W-030LFT).
_ID_GO_RE = re.compile(r"\bW-[A-Z0-9]{6}\b")


def _parse_equipo_from_case_id(case_id: str) -> str | None:
    """Extrae el token de equipo del case_id si sigue el formato CRM nuevo.

    Devuelve ``None`` si el case_id usa el formato heredado (``EV-2026-001``)
    o si el primer token no encaja con el patrón de equipo (dos letras +
    dos letras + dígitos, p. ej. ``BaRR3``, ``MaRS15``).
    """
    if not case_id:
        return None
    token = case_id.split(" - ", 1)[0].strip()
    return token if _EQUIPO_RE.match(token) else None


def _sanitize_filename_segment(s: str) -> str:
    """Sustituye los caracteres prohibidos en nombres de fichero Windows.

    Reemplaza ``/ \\ : * ? " < > |`` por espacios. No toca acentos ni
    caracteres válidos del case_id (`(`, `)`, `,`, `º`, etc.).
    """
    return _sanitize_filename_util(s, mode="segment")


def _parse_id_go_from_case_id(case_id: str) -> str | None:
    """Extrae el ID GO (``W-XXXXXX``) que el case_id lleva entre paréntesis.

    Tolera las dos convenciones que conviven en el Drive (``… 15 (W-02XOR7) -
    …`` y ``… 41 - (W-02MA0R) - …``) porque busca el patrón, no la posición.
    Devuelve ``None`` si no hay ID GO (caso real: ``(SIN REFERENCIA)``).
    """
    match = _ID_GO_RE.search(case_id or "")
    return match.group(0) if match else None


def _compose_informe_filename(case_id: str, id_go: str | None = None) -> str:
    """Compone el nombre del fichero del informe de viabilidad.

    Política (nombre con case_id completo decidido en la sesión 7 del
    2026-05-11; **acortado a solo el ID GO el 2026-07-28**):

    - Si ``case_id`` sigue el formato CRM nuevo (``<equipo> - <dirección>
      (<id_go>) - <sufijo>``, detectado por ``_parse_equipo_from_case_id``) y
      hay un ID GO resoluble: ``"Informe viabilidad - <id_go>.xlsx"``.
    - En cualquier otro caso (legacy ``EV-2026-001``, o formato nuevo sin ID GO
      como ``(SIN REFERENCIA)``): ``"_informe_viabilidad.xlsx"`` (fallback con
      underscore inicial). Mantiene ordenación arriba en el explorador y evita
      nombres con datos vacíos.

    El ``id_go`` explícito (el que ``ensure_case`` resuelve del frontmatter)
    manda sobre el que arrastra el case_id; si no es un ID GO válido se ignora
    y se cae al del case_id.

    **Por qué solo el ID GO y no el case_id completo.** El fichero vive en
    ``<CASOS>/<ciudad>/<case_id>/02_Analisis/``, así que repetir el case_id en
    el nombre no añade información y cuesta ~85 caracteres. Con el case_id
    completo el informe de W-02XOR7 llegaba a 269 caracteres de ruta y **Excel
    se negaba a abrirlo**: Office no es long-path aware y se rinde en 260
    aunque el sistema de ficheros admita más (``LongPathsEnabled=1`` y
    ``openpyxl`` sí abría el mismo fichero). El ID GO conserva la identidad del
    caso cuando el fichero viaja suelto (adjunto a un correo) por 8 caracteres.
    Se descartó truncar el case_id al hueco disponible porque el nombre
    dependería de dónde vive el caso, y el checkin mergea por ruta relativa:
    el mismo informe tendría un nombre en el Drive y otro en el checkout local.

    El nombre se sanea de caracteres prohibidos en Windows (``/ \\ : * ?
    " < > |``) por defensa, aunque un ID GO no debería contenerlos.
    """
    if _parse_equipo_from_case_id(case_id) is None:
        return "_informe_viabilidad.xlsx"
    for candidato in (id_go, _parse_id_go_from_case_id(case_id)):
        if candidato and _ID_GO_RE.fullmatch(candidato.strip()):
            return f"Informe viabilidad - {_sanitize_filename_segment(candidato.strip())}.xlsx"
    return "_informe_viabilidad.xlsx"


def _find_informe_existente(analisis_dir: Path) -> Path | None:
    """Localiza un informe de viabilidad ya presente, con cualquiera de los
    nombres históricos.

    Necesario tras el acortamiento del 2026-07-28: los casos abiertos antes
    llevan el case_id completo en el nombre, y comparar solo contra el destino
    nuevo dejaría una segunda plantilla en blanco al lado del informe que el
    abogado ya ha trabajado. Reconoce el nombre nuevo, el largo legacy, el
    fallback ``_informe_viabilidad.xlsx`` y los puestos a mano en el Drive
    (hay uno en mayúsculas), comparando en minúsculas y con ``_`` como espacio.

    Excluye ``Informe viabilidad LLM - …`` a propósito: es el artefacto
    paralelo que genera ``render_informe.py`` y no sustituye al informe humano.
    """
    if not analisis_dir.is_dir():
        return None
    for path in sorted(analisis_dir.iterdir()):
        if path.suffix.lower() != ".xlsx" or not path.is_file():
            continue
        stem = path.stem.lower().replace("_", " ").strip()
        if stem.startswith("informe viabilidad") and "llm" not in stem:
            return path
    return None


def _avisar_si_ruta_larga(path: Path, *, limite: int = RUTA_OFFICE_MAX) -> bool:
    """Avisa (no aborta) si la ruta supera el presupuesto que tolera Office.

    Guardarraíl de defensa en profundidad: el nombre corto resuelve el caso
    conocido, pero una carpeta de caso bautizada con una dirección muy larga
    podría volver a pasarse. Devuelve ``True`` si ha avisado.
    """
    largo = len(str(path))
    if largo <= limite:
        return False
    logger.warning(
        "Ruta de %d caracteres (presupuesto %d) para %s: Excel no abre "
        "ficheros cuya ruta completa roza los 260 aunque el sistema de "
        "ficheros lo admita. Acorta el nombre de la carpeta del caso.",
        largo,
        limite,
        path.name,
    )
    return True


def _ensure_crm_tree_dirs(case_dir: Path) -> None:
    """Crea solo la base ``00_Input/05_CRM/`` (D7 — andamiaje *lazy*).

    Tras el aplanado a buckets planos (D5/D6, reorg 2026-06-10) ya NO se
    pre-crea el árbol profundo del CRM: los buckets se materializan al
    escribir el primer documento. Tanto el pull (``pull_expediente_v2`` hace
    ``dest.parent.mkdir(parents=True, exist_ok=True)`` antes de cada
    ``write_bytes``) como el intake manual (``save_file_crm_branch``) crean su
    bucket on-write. Esto evita el bosque de ~20 carpetas vacías del
    andamiaje eager anterior (D1) y los problemas de longitud de ruta en
    Windows sobre el Drive.
    """
    base = case_dir / "00_Input" / CRM_SUBDIR
    base.mkdir(parents=True, exist_ok=True)


def _copy_plantilla(origen: Path, destino: Path) -> bool:
    """Copia una plantilla al destino con idempotencia estricta.

    - Si ``destino`` ya existe → NO sobrescribe (preserva trabajo del
      abogado) y devuelve ``False``.
    - Si ``origen`` no existe → log warning y devuelve ``False`` (no
      aborta ensure_case; podría faltar tras un checkout parcial o un
      mount intermitente de la unidad de red).
    - En otro caso, copia con ``shutil.copy2`` (preserva metadatos) y
      devuelve ``True``.
    """
    if destino.exists():
        return False
    if not origen.exists():
        logger.warning("Plantilla no encontrada, se omite copia: %s", origen)
        return False
    destino.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(origen, destino)
    return True


def _find_label_row(
    ws: Any,
    label: str,
    *,
    column: int = 2,
    max_rows: int = 200,
) -> int | None:
    """Localiza la primera fila cuya celda en ``column`` matchea ``label``.

    Las celdas concretas donde el render escribe REF/FECHA no son
    estables (dependen del orden de los bloques en el YAML). Localizar
    por etiqueta es resistente a cambios futuros del template.
    """
    for row in range(1, max_rows + 1):
        value = ws.cell(row=row, column=column).value
        if value is not None and str(value).strip() == label:
            return row
    return None


def _prerellenar_informe(
    informe_path: Path,
    *,
    equipo: str | None,
    direccion: str | None,
    id_go: str | None,
) -> None:
    """Pre-rellena REF y FECHA en el informe de viabilidad recién copiado.

    Localiza ``REF`` y ``FECHA`` en columna B de la hoja ``OPERACION`` y
    escribe el valor en la columna C de la misma fila. REF se rellena
    SOLO si los tres componentes están presentes (D-7a-2) — sin
    placeholders. FECHA siempre se rellena con la fecha actual.

    Silencioso ante errores de openpyxl: un informe corrupto no debe
    abortar la creación del caso.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl no disponible; se omite pre-relleno del informe")
        return

    try:
        wb = load_workbook(informe_path)
    except Exception as exc:
        logger.warning("No se pudo abrir informe para pre-relleno (%s): %s", informe_path, exc)
        return

    if "OPERACION" not in wb.sheetnames:
        logger.warning("Hoja 'OPERACION' no encontrada en %s", informe_path)
        return
    ws = wb["OPERACION"]

    ref_row = _find_label_row(ws, "REF")
    if ref_row is not None and equipo and direccion and id_go:
        ws.cell(row=ref_row, column=3, value=f"{equipo} - {direccion} ({id_go})")

    fecha_row = _find_label_row(ws, "FECHA")
    if fecha_row is not None:
        cell = ws.cell(row=fecha_row, column=3, value=date.today())
        cell.number_format = "dd/mm/yyyy"

    try:
        wb.save(informe_path)
    except Exception as exc:
        logger.warning("No se pudo guardar informe tras pre-relleno (%s): %s", informe_path, exc)
