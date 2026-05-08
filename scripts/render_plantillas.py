"""Renderiza los XLSX de plantilla a partir de los YAML canónicos.

Los YAML de ``data/_plantillas/`` son la fuente única. Este script genera
los XLSX correspondientes que el equipo abre con Excel:

- ``cuestionario_viabilidad.yaml`` → ``cuestionario_viabilidad.xlsx`` (11 hojas).
- ``ficha_operacion.yaml`` → ``ficha_operacion.xlsx`` (1 hoja `OPERACION`).

Workflow recomendado:

    cd "G:\\Unidades compartidas\\DESPACHO - PRODUCCION\\Base datos expedientes"
    python -m scripts.render_plantillas all

El XLSX generado lleva una hoja oculta ``_meta`` con la versión del YAML
y un hash corto del fichero fuente — útil para detectar plantillas
desactualizadas en producción.
"""

from __future__ import annotations

import hashlib
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import typer
import yaml
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet


PROJECT_ROOT = Path(__file__).resolve().parent.parent
PLANTILLAS_DIR = PROJECT_ROOT / "data" / "_plantillas"


# ---------------------------------------------------------------------------
# Estilos comunes
# ---------------------------------------------------------------------------

FILL_HEADER = PatternFill("solid", fgColor="DEEAF6")
FILL_PRERELLENO = PatternFill("solid", fgColor="F2F2F2")
FILL_TOTAL = PatternFill("solid", fgColor="D9E2F3")

FONT_TITLE = Font(name="Calibri", size=14, bold=True, color="1F4E79")
FONT_HEADER = Font(name="Calibri", size=10, bold=True)
FONT_BODY = Font(name="Calibri", size=10)
FONT_SMALL = Font(name="Calibri", size=9)
FONT_TOTAL = Font(name="Calibri", size=11, bold=True, color="1F4E79")

THIN = Side(border_style="thin", color="9BB1D8")
BORDER_CELL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Helpers comunes
# ---------------------------------------------------------------------------

class _StrictBoolLoader(yaml.SafeLoader):
    """SafeLoader que solo interpreta ``true``/``false`` como bool.

    PyYAML por defecto sigue YAML 1.1 e interpreta ``yes``/``no``/``on``/``off``
    como booleanos, lo que choca con el uso natural de ``si``/``no`` como
    valores de respuesta o de opciones de enum en nuestros YAML.
    """


# Reemplaza los resolvers de bool para que NO matcheen si/no/yes/on/off.
_StrictBoolLoader.yaml_implicit_resolvers = {
    k: [(tag, regexp) for tag, regexp in resolvers
        if tag != "tag:yaml.org,2002:bool"]
    for k, resolvers in _StrictBoolLoader.yaml_implicit_resolvers.items()
}
_StrictBoolLoader.add_implicit_resolver(
    "tag:yaml.org,2002:bool",
    re.compile(r"^(?:true|True|TRUE|false|False|FALSE)$"),
    list("tTfF"),
)


def load_yaml(path: Path) -> dict:
    with open(path, encoding="utf-8") as f:
        return yaml.load(f, Loader=_StrictBoolLoader)


def file_hash_short(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()[:16]


def write_meta_sheet(wb: Workbook, yaml_path: Path, data: dict) -> None:
    """Hoja oculta `_meta` con info de la generación."""
    ws = wb.create_sheet(title="_meta")
    ws.append(["version", data.get("version")])
    ws.append(["ultima_actualizacion", str(data.get("ultima_actualizacion") or "")])
    ws.append(["referencia_origen", data.get("referencia_origen") or ""])
    ws.append(["yaml_sha256_16", file_hash_short(yaml_path)])
    ws.append(["generado", datetime.now().isoformat(timespec="seconds")])
    ws.sheet_state = "hidden"


def safe_sheet_title(title: str) -> str:
    """Excel limita el nombre de hoja a 31 chars y prohíbe `/\\?*[]:`."""
    clean = title
    for bad in ("/", "\\", "?", "*", "[", "]", ":"):
        clean = clean.replace(bad, " ")
    return clean.strip()[:31]


# ---------------------------------------------------------------------------
# Cuestionario — render
# ---------------------------------------------------------------------------

CUESTIONARIO_COLUMNS: list[tuple[str, int]] = [
    ("ID",                            10),
    ("Pregunta",                      55),
    ("Objetivo probatorio",           40),
    ("Tipo",                          14),
    ("Fuente probable",               22),
    ("Respuesta",                     22),
    ("Cita literal del documento",    40),
    ("Confianza LLM",                 14),
    ("Validado",                      12),
]


def render_cuestionario(yaml_path: Path, xlsx_path: Path) -> None:
    data = load_yaml(yaml_path)
    wb = Workbook()
    wb.remove(wb.active)

    secciones = data.get("secciones", [])
    for sec in secciones:
        ws = wb.create_sheet(title=safe_sheet_title(sec["titulo"]))
        _render_seccion(ws, sec)

    write_meta_sheet(wb, yaml_path, data)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def _render_seccion(ws: Worksheet, sec: dict) -> None:
    n_cols = len(CUESTIONARIO_COLUMNS)

    # Fila 1: título de la sección
    titulo_cell = ws.cell(row=1, column=1, value=sec["titulo"])
    titulo_cell.font = FONT_TITLE
    titulo_cell.fill = FILL_HEADER
    titulo_cell.alignment = ALIGN_LEFT_CENTER
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.row_dimensions[1].height = 26

    # Fila 2: descripción (si la hay)
    desc = (sec.get("descripcion") or "").strip()
    if desc:
        desc_cell = ws.cell(row=2, column=1, value=desc)
        desc_cell.font = FONT_SMALL
        desc_cell.alignment = ALIGN_LEFT_TOP
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        ws.row_dimensions[2].height = max(20, 12 + len(desc) // 6)

    # Fila 3: cabeceras de la tabla
    for idx, (label, width) in enumerate(CUESTIONARIO_COLUMNS, start=1):
        c = ws.cell(row=3, column=idx, value=label)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_CELL
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.row_dimensions[3].height = 22

    # Filas 4+: preguntas
    preguntas = sec.get("preguntas", [])
    for offset, preg in enumerate(preguntas):
        r = 4 + offset
        _render_pregunta(ws, r, preg)

    ws.freeze_panes = "B4"
    ws.sheet_view.showGridLines = False


def _render_pregunta(ws: Worksheet, row: int, preg: dict) -> None:
    fuente = preg.get("fuente_probable") or []
    fuente_str = ", ".join(fuente) if isinstance(fuente, list) else str(fuente)
    obj = (preg.get("objetivo_probatorio") or "").strip()
    tipo = preg.get("tipo_respuesta", "")

    ws.cell(row=row, column=1, value=preg["id"]).font = FONT_BODY
    ws.cell(row=row, column=2, value=preg["texto"]).font = FONT_BODY
    ws.cell(row=row, column=3, value=obj).font = FONT_SMALL
    ws.cell(row=row, column=4, value=tipo).font = FONT_BODY
    ws.cell(row=row, column=5, value=fuente_str).font = FONT_SMALL
    ws.cell(row=row, column=6, value="").font = FONT_BODY
    ws.cell(row=row, column=7, value="").font = FONT_SMALL
    ws.cell(row=row, column=8, value="").font = FONT_BODY
    ws.cell(row=row, column=9, value="").font = FONT_BODY

    # Estilos: A-E con fondo gris claro (visualmente "read-only"),
    #         F-I editables, todas con borde y alineación tope-izq.
    for col in range(1, 10):
        c = ws.cell(row=row, column=col)
        c.alignment = ALIGN_LEFT_TOP
        c.border = BORDER_CELL
        if col <= 5:
            c.fill = FILL_PRERELLENO

    # Altura fila proporcional al texto
    text_len = len(preg["texto"]) + len(obj)
    ws.row_dimensions[row].height = max(28, min(80, 18 + text_len // 10))

    _add_response_validation(ws, row, col=6, preg=preg)
    _add_validado_validation(ws, row, col=9)


def _add_response_validation(ws: Worksheet, row: int, col: int, preg: dict) -> None:
    """Validación de datos en la celda de Respuesta según el tipo de pregunta."""
    addr = f"{get_column_letter(col)}{row}"
    tipo = preg.get("tipo_respuesta", "texto_libre")

    dv: DataValidation | None = None
    if tipo == "boolean":
        dv = DataValidation(type="list", formula1='"Sí,No,Pendiente"', allow_blank=True)
    elif tipo == "enum":
        opciones = preg.get("opciones") or []
        if opciones:
            joined = ",".join(str(o) for o in opciones)
            if len(joined) <= 250:
                dv = DataValidation(type="list", formula1=f'"{joined}"', allow_blank=True)
    elif tipo == "numero":
        dv = DataValidation(type="decimal", operator="greaterThanOrEqual",
                            formula1="0", allow_blank=True)
    elif tipo == "fecha":
        ws.cell(row=row, column=col).number_format = "dd/mm/yyyy"
        # Sin validación estricta: el usuario puede escribir "N/A" como texto.

    if dv is not None:
        dv.add(addr)
        ws.add_data_validation(dv)


def _add_validado_validation(ws: Worksheet, row: int, col: int) -> None:
    addr = f"{get_column_letter(col)}{row}"
    dv = DataValidation(type="list", formula1='"Sí,No"', allow_blank=True)
    dv.add(addr)
    ws.add_data_validation(dv)


# ---------------------------------------------------------------------------
# Ficha — render
# ---------------------------------------------------------------------------

# Layout de la ficha (columnas de Excel):
#   A: padding
#   B: etiqueta
#   C: valor
#   D: fecha (solo en datos_operacion)
#   E: observaciones (solo en datos_operacion)
#   F: padding
#
# Los bloques no-tabla usan B (etiqueta) + C (valor) y dejan D, E vacías.
# El bloque datos_operacion usa B/C/D/E como cabeceras de tabla.

FICHA_WIDTHS = {"A": 2, "B": 45, "C": 22, "D": 14, "E": 50, "F": 2}


def render_ficha(yaml_path: Path, xlsx_path: Path) -> None:
    data = load_yaml(yaml_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "OPERACION"
    ws.sheet_view.showGridLines = False

    for col, w in FICHA_WIDTHS.items():
        ws.column_dimensions[col].width = w

    campo_celdas: dict[str, str] = {}
    row = 1
    for bloque in data.get("bloques", []):
        row = _render_bloque_ficha(ws, row, bloque, campo_celdas)
        row += 1  # separación entre bloques

    _resolve_ficha_formulas(ws, campo_celdas, data)

    write_meta_sheet(wb, yaml_path, data)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def _render_bloque_ficha(
    ws: Worksheet,
    row: int,
    bloque: dict,
    campo_celdas: dict[str, str],
) -> int:
    """Renderiza un bloque de la ficha. Devuelve la fila siguiente libre."""
    titulo = bloque.get("titulo", "")
    bid = bloque.get("id")

    # Cabecera del bloque (B-E merged)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    cell = ws.cell(row=row, column=2, value=titulo)
    cell.font = FONT_TITLE
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_CENTER
    ws.row_dimensions[row].height = 24
    row += 1

    desc = (bloque.get("descripcion") or "").strip()
    if desc:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        c = ws.cell(row=row, column=2, value=desc)
        c.font = FONT_SMALL
        c.alignment = ALIGN_LEFT_TOP
        ws.row_dimensions[row].height = max(20, 14 + len(desc) // 5)
        row += 1

    if bid == "datos_operacion":
        return _render_datos_operacion(ws, row, bloque, campo_celdas)
    if bid == "viabilidad":
        return _render_viabilidad(ws, row, bloque, campo_celdas)
    return _render_bloque_dos_columnas(ws, row, bloque, campo_celdas)


def _render_bloque_dos_columnas(
    ws: Worksheet,
    row: int,
    bloque: dict,
    campo_celdas: dict[str, str],
) -> int:
    """Bloques cabecera, equipo, contexto, importes, actividades."""
    for campo in bloque.get("campos", []):
        cid = campo.get("id")
        etiqueta = campo.get("etiqueta", "")
        tipo = campo.get("tipo", "texto_libre")
        valor_default = campo.get("valor_default")

        et_cell = ws.cell(row=row, column=2, value=etiqueta)
        et_cell.font = FONT_HEADER
        et_cell.alignment = ALIGN_LEFT_TOP
        et_cell.border = BORDER_CELL

        v_cell = ws.cell(row=row, column=3)
        v_cell.alignment = ALIGN_LEFT_TOP
        v_cell.border = BORDER_CELL

        if tipo == "formula":
            # Fórmula resuelta en una segunda pasada (resolve_formulas).
            v_cell.value = f"<<FORMULA:{cid}>>"
            v_cell.number_format = "#,##0.00"
        elif valor_default is not None:
            v_cell.value = valor_default
            if tipo == "numero":
                v_cell.number_format = "#,##0.00"
        else:
            v_cell.value = ""

        if tipo == "fecha":
            v_cell.number_format = "dd/mm/yyyy"
        elif tipo == "numero":
            v_cell.number_format = "#,##0.00"

        if tipo == "texto_libre_multiline":
            ws.row_dimensions[row].height = 50
        else:
            ws.row_dimensions[row].height = 20

        # D y E quedan vacías (con borde para que la cuadrícula del bloque sea consistente)
        for ec in (4, 5):
            ws.cell(row=row, column=ec).border = BORDER_CELL

        campo_celdas[cid] = f"C{row}"
        row += 1

    return row


def _render_viabilidad(
    ws: Worksheet,
    row: int,
    bloque: dict,
    campo_celdas: dict[str, str],
) -> int:
    """Tres enums con formato condicional verde/amarillo/rojo/pendiente."""
    first_value_cell = None
    last_value_cell = None

    for campo in bloque.get("campos", []):
        cid = campo.get("id")
        etiqueta = campo.get("etiqueta", "")
        opciones = campo.get("opciones") or []
        valor_default = campo.get("valor_default", "")

        et_cell = ws.cell(row=row, column=2, value=etiqueta)
        et_cell.font = FONT_HEADER
        et_cell.alignment = ALIGN_LEFT_TOP
        et_cell.border = BORDER_CELL

        v_cell = ws.cell(row=row, column=3, value=valor_default)
        v_cell.alignment = ALIGN_CENTER
        v_cell.border = BORDER_CELL

        if opciones:
            joined = ",".join(str(o) for o in opciones)
            dv = DataValidation(type="list", formula1=f'"{joined}"', allow_blank=True)
            dv.add(f"C{row}")
            ws.add_data_validation(dv)

        # D y E vacías con borde
        for ec in (4, 5):
            ws.cell(row=row, column=ec).border = BORDER_CELL

        ref = f"C{row}"
        first_value_cell = first_value_cell or ref
        last_value_cell = ref
        campo_celdas[cid] = ref
        ws.row_dimensions[row].height = 22
        row += 1

    fc_rules = bloque.get("formato_condicional") or []
    if fc_rules and first_value_cell and last_value_cell:
        rng = f"{first_value_cell}:{last_value_cell}"
        for fc in fc_rules:
            valor = fc["valor"]
            fondo = fc["fondo"]
            texto = fc["texto"]
            ws.conditional_formatting.add(
                rng,
                CellIsRule(
                    operator="equal",
                    formula=[f'"{valor}"'],
                    stopIfTrue=False,
                    fill=PatternFill("solid", fgColor=fondo),
                    font=Font(color=texto, bold=True),
                ),
            )

    return row


def _render_datos_operacion(
    ws: Worksheet,
    row: int,
    bloque: dict,
    campo_celdas: dict[str, str],
) -> int:
    """Tabla 4-columnas: HITO | SCORE | FECHA | OBSERVACIONES + fila TOTAL."""
    headers = [("HITO", 2), ("SCORE", 3), ("FECHA", 4), ("OBSERVACIONES", 5)]
    for label, col in headers:
        c = ws.cell(row=row, column=col, value=label)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_CELL
    ws.row_dimensions[row].height = 22
    row += 1

    primera_fila_score = row

    for hito in bloque.get("hitos", []):
        hid = hito["id"]

        # B — etiqueta
        et = ws.cell(row=row, column=2, value=hito.get("etiqueta", ""))
        et.font = FONT_BODY
        et.alignment = ALIGN_LEFT_TOP
        et.border = BORDER_CELL

        # C — score
        sc = ws.cell(row=row, column=3)
        sc.alignment = ALIGN_CENTER
        sc.border = BORDER_CELL
        sc.number_format = "0"

        score_block = hito.get("score") or {}
        regla = score_block.get("regla") or {}
        if regla.get("formula_excel"):
            # Cuantía — fórmula que se resuelve en _resolve_ficha_formulas.
            sc.value = f"<<FORMULA_SCORE:{hid}>>"
        else:
            sc.value = ""
            dv = DataValidation(
                type="list", formula1='"0,1,2,3,N/A"', allow_blank=True,
            )
            dv.add(f"C{row}")
            ws.add_data_validation(dv)

        # D — fecha
        fc = ws.cell(row=row, column=4)
        fc.alignment = ALIGN_CENTER
        fc.border = BORDER_CELL
        if hito.get("fecha"):
            fc.number_format = "dd/mm/yyyy"

        # E — observaciones
        oc = ws.cell(row=row, column=5)
        oc.alignment = ALIGN_LEFT_TOP
        oc.border = BORDER_CELL

        ws.row_dimensions[row].height = 26
        campo_celdas[f"hito_{hid}_score"] = f"C{row}"
        row += 1

    ultima_fila_score = row - 1

    # Fila TOTAL
    et = ws.cell(row=row, column=2, value="TOTAL")
    et.font = FONT_TOTAL
    et.fill = FILL_TOTAL
    et.alignment = ALIGN_LEFT_CENTER
    et.border = BORDER_CELL

    total_cell = ws.cell(
        row=row, column=3,
        value=f"=SUM(C{primera_fila_score}:C{ultima_fila_score})",
    )
    total_cell.font = FONT_TOTAL
    total_cell.fill = FILL_TOTAL
    total_cell.alignment = ALIGN_CENTER
    total_cell.border = BORDER_CELL
    total_cell.number_format = "0"

    for ec in (4, 5):
        c = ws.cell(row=row, column=ec)
        c.fill = FILL_TOTAL
        c.border = BORDER_CELL

    ws.row_dimensions[row].height = 24
    return row + 1


def _resolve_ficha_formulas(
    ws: Worksheet, campo_celdas: dict[str, str], data: dict,
) -> None:
    """Sustituye placeholders <<FORMULA:...>> y <<FORMULA_SCORE:...>>."""
    bloques = {b["id"]: b for b in data.get("bloques", [])}

    # Bloque importes — fórmulas en cascada
    importes = bloques.get("importes", {})
    for campo in importes.get("campos", []):
        if campo.get("tipo") != "formula":
            continue
        cid = campo["id"]
        cell_ref = campo_celdas.get(cid)
        if not cell_ref:
            continue
        template = campo.get("formula_excel", "")
        ws[cell_ref] = _replace_placeholders(template, campo_celdas)
        ws[cell_ref].number_format = "#,##0.00"

    # Datos operación — fórmula CUANTÍA
    datos_op = bloques.get("datos_operacion", {})
    for hito in datos_op.get("hitos", []):
        regla = (hito.get("score") or {}).get("regla") or {}
        template = regla.get("formula_excel")
        if not template:
            continue
        cell_ref = campo_celdas.get(f"hito_{hito['id']}_score")
        if not cell_ref:
            continue
        ws[cell_ref] = _replace_placeholders(template, campo_celdas)


_PLACEHOLDER_KEYS = {
    "PRECIO":           "precio",
    "PCT_HONORARIOS":   "porcentaje_honorarios",
    "TOTAL_HONORARIOS": "total_honorarios",
    "PAGOS_PARCIALES":  "pagos_parciales",
    "TOTAL_DEUDA":      "total_deuda",
    "PROPUESTA_PAGO":   "propuesta_pago",
}


def _replace_placeholders(template: str, campo_celdas: dict[str, str]) -> str:
    """Sustituye nombres simbólicos (PRECIO, TOTAL_DEUDA…) por celdas reales."""
    formula = template
    for key in sorted(_PLACEHOLDER_KEYS, key=len, reverse=True):
        cell_ref = campo_celdas.get(_PLACEHOLDER_KEYS[key], "")
        formula = formula.replace(key, cell_ref)
    return formula


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

app = typer.Typer(add_completion=False, no_args_is_help=True, pretty_exceptions_enable=False)


@app.command()
def cuestionario() -> None:
    """Genera cuestionario_viabilidad.xlsx desde el YAML."""
    yaml_p = PLANTILLAS_DIR / "cuestionario_viabilidad.yaml"
    xlsx_p = PLANTILLAS_DIR / "cuestionario_viabilidad.xlsx"
    if not yaml_p.exists():
        raise typer.Exit(code=1)
    render_cuestionario(yaml_p, xlsx_p)
    typer.echo(f"OK  {xlsx_p}")


@app.command()
def ficha() -> None:
    """Genera ficha_operacion.xlsx desde el YAML."""
    yaml_p = PLANTILLAS_DIR / "ficha_operacion.yaml"
    xlsx_p = PLANTILLAS_DIR / "ficha_operacion.xlsx"
    if not yaml_p.exists():
        raise typer.Exit(code=1)
    render_ficha(yaml_p, xlsx_p)
    typer.echo(f"OK  {xlsx_p}")


@app.command(name="all")
def render_all() -> None:
    """Genera ambos XLSX."""
    cuestionario()
    ficha()


if __name__ == "__main__":
    app()
