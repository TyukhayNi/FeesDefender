#!/usr/bin/env python3
"""
render_informe.py — Rellena la plantilla del Informe de Viabilidad (E&V / FeesDefender)
a partir de un JSON de datos extraídos en la 1ª pasada documental (Skill A).

PRINCIPIO: parte SIEMPRE de assets/plantilla_informe_viabilidad.xlsx (trae formato,
fórmulas, semáforo condicional, validaciones y protección). Este script SOLO escribe
valores; nunca regenera el formato.

Lo que ESCRIBE: cabecera, equipo, observaciones, MOTIVOS (si procede), importes
(inputs, conservando fórmulas), 14 hitos (score+fecha), actividades, hoja PREGUNTAS
(RESPUESTA/CITA/CONFIANZA/¿PENDIENTE?), hoja AVISOS LLM y la 1ª entrada de BITACORA.

Lo que NUNCA toca: VIABILIDAD (E21/E22, siempre en blanco en el pre-relleno),
el recuadro ejecutivo (B48, lo escribe la Skill B), las columnas fijas del cuestionario
ni las NOTAS LETRADO (las rellena el abogado).

SALIDA: fichero paralelo. NUNCA sobrescribe el informe humano.

Uso:
    python render_informe.py datos.json --salida "Informe viabilidad LLM - <case_id>.xlsx"
    python render_informe.py datos.json            # deriva el nombre de case_id en el JSON
"""
import argparse, json, os, sys, shutil
from datetime import datetime

try:
    import openpyxl
    from openpyxl.cell.cell import MergedCell
except ImportError:
    sys.exit("Falta openpyxl: pip install openpyxl --break-system-packages")

HERE = os.path.dirname(os.path.abspath(__file__))
PLANTILLA = os.path.join(HERE, "..", "assets", "plantilla_informe_viabilidad.xlsx")

# Hito canónico -> fila en INFORMACION (score en col F, fecha en col H)
HITO_ROWS = {
    "CUANTIA": 25, "ENCARGO": 26, "IDENT_PROPIETARIO": 27, "TITULARIDAD": 28,
    "HOJA_VISITA": 29, "OFERTA": 30, "IDENT_BUSCADOR": 31, "ARRAS_ARRENDAMIENTO": 32,
    "RECON_HON_ARRAS": 33, "ESCRITURA": 34, "RECON_HON_ESCRITURA": 35,
    "RECLAMACION_JURIDICO": 36, "RESPUESTA_RECLAMACION": 37, "OFERTA_VINCULANTE_CONFIDENCIAL": 38,
}
# Aliases tolerantes (por si el JSON usa el rótulo de pantalla)
HITO_ALIASES = {
    "CUANTÍA": "CUANTIA",
    "IDENTIFICACIÓN PROPIETARIO": "IDENT_PROPIETARIO",
    "IDENTIFICACIÓN - BUSCADOR": "IDENT_BUSCADOR", "IDENTIFICACIÓN BUSCADOR": "IDENT_BUSCADOR",
    "HOJA DE VISITA": "HOJA_VISITA", "HOJA VISITA": "HOJA_VISITA",
    "ARRAS/ARRENDAMIENTO": "ARRAS_ARRENDAMIENTO", "ARRAS / ARRENDAMIENTO": "ARRAS_ARRENDAMIENTO",
    "RECON. HON. — ARRAS": "RECON_HON_ARRAS", "RECONOCIMIENTO HONORARIOS -ARRAS": "RECON_HON_ARRAS",
    "RECON. HON. — ESCRITURA": "RECON_HON_ESCRITURA", "RECONOCIMIENTO HONORARIOS - ESCRITURA": "RECON_HON_ESCRITURA",
    "RECLAMACIÓN JURÍDICO": "RECLAMACION_JURIDICO", "RECLAMACIÓN JURIDICO": "RECLAMACION_JURIDICO",
    "RESPUESTA A LA RECLAMACIÓN": "RESPUESTA_RECLAMACION", "RESPUESTA RECLAMACIÓN": "RESPUESTA_RECLAMACION",
    "OFERTA VINCULANTE CONFIDENCIAL": "OFERTA_VINCULANTE_CONFIDENCIAL",
}

EQUIPO_CELLS = {
    "director_captador": "E6", "asesor_captador": "E7",
    "director_buscador": "E8", "asesor_buscador": "E9",
}


def warn(msg):
    print(f"  [aviso] {msg}", file=sys.stderr)


def set_cell(ws, coord, value):
    """Escribe en una celda; si es una celda combinada no-ancla, avisa y salta."""
    cell = ws[coord]
    if isinstance(cell, MergedCell):
        warn(f"{ws.title}!{coord} es celda combinada no-ancla; no se escribe.")
        return
    cell.value = value


def build_id_row_map(ws_preg):
    """ID de pregunta (col C) -> nº de fila, escaneando la hoja PREGUNTAS."""
    m = {}
    for r in range(5, ws_preg.max_row + 1):
        idq = ws_preg.cell(r, 3).value
        if idq:
            m[str(idq).strip()] = r
    return m


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("datos_json", help="JSON con los datos extraídos en la 1ª pasada")
    ap.add_argument("--salida", help="Ruta del .xlsx de salida")
    ap.add_argument("--plantilla", default=PLANTILLA, help="Plantilla base (por defecto, la de assets/)")
    args = ap.parse_args()

    with open(args.datos_json, encoding="utf-8") as f:
        d = json.load(f)

    salida = args.salida
    if not salida:
        cid = d.get("case_id", "SIN_ID")
        salida = f"Informe viabilidad LLM - {cid}.xlsx"

    if os.path.abspath(salida) == os.path.abspath(args.plantilla):
        sys.exit("ERROR: la salida no puede ser la propia plantilla.")
    if os.path.exists(salida):
        sys.exit(f"ERROR: '{salida}' ya existe. Renombra o borra antes (la skill nunca sobrescribe).")

    shutil.copy(args.plantilla, salida)
    wb = openpyxl.load_workbook(salida)
    inf = wb["INFORMACION"]; preg = wb["PREGUNTAS"]
    av = wb["AVISOS LLM"]; bit = wb["BITACORA"]

    # --- Cabecera ---
    set_cell(inf, "E4", d.get("fecha") or datetime.today().strftime("%d/%m/%Y"))
    if d.get("ref"):
        set_cell(inf, "E5", d["ref"])

    # --- Equipo comercial (Apellido, Nombre) ---
    for k, coord in EQUIPO_CELLS.items():
        v = (d.get("equipo") or {}).get(k)
        if v:
            set_cell(inf, coord, v)

    # --- Observaciones (etiqueta tipo de caso) ---
    if d.get("observaciones"):
        set_cell(inf, "E11", d["observaciones"])

    # --- MOTIVOS DE IMPAGO (H12): vacío por defecto; solo si consta postura del deudor ---
    motivos = (d.get("motivos_impago") or "").strip()
    if motivos:
        set_cell(inf, "H12", motivos.upper())

    # --- Importes (inputs; las fórmulas H14/H16/H18 ya están en la plantilla) ---
    imp = d.get("importes") or {}
    if imp.get("precio") is not None:           set_cell(inf, "H13", imp["precio"])
    set_cell(inf, "E14", imp.get("pct_honorarios", 5))          # default 5
    set_cell(inf, "H15", imp.get("pagos_parciales", 0))         # default 0
    if imp.get("propuesta_pago") is not None:   set_cell(inf, "H17", imp["propuesta_pago"])

    # --- 14 hitos (score + fecha). 'pendiente' o None => celda vacía, NUNCA 0 por inferencia ---
    hitos = d.get("hitos") or {}
    for raw_key, val in hitos.items():
        key = HITO_ALIASES.get(raw_key.strip(), raw_key.strip())
        row = HITO_ROWS.get(key)
        if not row:
            warn(f"hito desconocido '{raw_key}' — se ignora.")
            continue
        score = val.get("score") if isinstance(val, dict) else val
        fecha = val.get("fecha") if isinstance(val, dict) else None
        if score in (None, "pendiente", "PENDIENTE", ""):
            pass  # se deja vacío
        else:
            set_cell(inf, f"F{row}", score)  # 1/2/3, 0, o "N/A"
        if fecha:
            set_cell(inf, f"H{row}", fecha)

    # --- Actividades ---
    act = d.get("actividades") or {}
    for k, coord in {"exposes_propiedad": "F42", "visitas_propiedad": "F43",
                     "exposes_buscador": "F44", "visitas_buscador": "F45"}.items():
        if act.get(k) is not None:
            set_cell(inf, coord, act[k])

    # --- VIABILIDAD: NO se toca (E21/E22 quedan en blanco en el pre-relleno) ---

    # --- PREGUNTAS: columnas del LLM ---
    id_row = build_id_row_map(preg)
    for qid, ans in (d.get("preguntas") or {}).items():
        r = id_row.get(qid)
        if not r:
            warn(f"pregunta '{qid}' no está en la plantilla — se ignora.")
            continue
        if ans.get("respuesta") is not None: preg.cell(r, 9).value = ans["respuesta"]   # I
        if ans.get("cita") is not None:      preg.cell(r, 10).value = ans["cita"]        # J
        if ans.get("confianza"):             preg.cell(r, 11).value = ans["confianza"]   # K alta/media/baja
        # ¿PENDIENTE ENTREVISTA? (M): default 'no' si hay respuesta documental; 'sí' si no
        pend = ans.get("pendiente")
        if pend is None:
            pend = "no" if ans.get("respuesta") not in (None, "", "pendiente") else "sí"
        preg.cell(r, 13).value = pend

    # --- AVISOS LLM (capa de trabajo) ---
    avisos = d.get("avisos") or []
    for i, a in enumerate(avisos):
        r = 5 + i
        av.cell(r, 2).value = a.get("n", i + 1)                       # B Nº
        av.cell(r, 3).value = a.get("tipo")                          # C TIPO
        av.cell(r, 4).value = a.get("aviso")                         # D AVISO
        av.cell(r, 5).value = a.get("impacto")                       # E IMPACTO
        av.cell(r, 6).value = a.get("fuente")                        # F FUENTE/RASTRO
        av.cell(r, 7).value = a.get("severidad")                     # G alta/media/baja
        av.cell(r, 8).value = a.get("accion")                        # H ACCIÓN
        av.cell(r, 9).value = a.get("sube", "no")                    # I ¿SUBE? (lo decide el abogado)
        av.cell(r, 10).value = a.get("estado", "abierto")            # J ESTADO

    # --- BITÁCORA: primera entrada del ciclo (pre-relleno) ---
    if d.get("bitacora_inicial", True):
        fecha = d.get("fecha") or datetime.today().strftime("%d/%m/%Y")
        bit.cell(5, 2).value = fecha
        bit.cell(5, 3).value = "Pre-relleno documental"
        bit.cell(5, 4).value = "Administración / LLM → expediente"
        bit.cell(5, 5).value = ("Carga inicial de los hitos desde 00_Input. "
                                "Viabilidad en PENDIENTE. Testificales pendientes de entrevista.")
        bit.cell(5, 6).value = "—"
        bit.cell(5, 7).value = "Pendiente entrevista"

    wb.save(salida)
    print(f"OK · Informe generado: {salida}")
    print("   VIABILIDAD en blanco · recuadro ejecutivo vacío (los completa la Skill B / el abogado).")


if __name__ == "__main__":
    main()
