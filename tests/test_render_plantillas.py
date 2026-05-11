"""Tests dedicados v2 — paso 8 del refactor intake v2.

Módulo bajo test: ``scripts.render_plantillas`` (YAML→XLSX para las
plantillas de viabilidad).

Estrategia (cerrada en la sesión 2026-05-11):

- Smoke estructural: el XLSX se genera, las hojas críticas existen, la
  hoja ``_meta`` queda oculta y contiene el hash del YAML fuente.
- Contrato del ``_StrictBoolLoader``: ``si``/``no``/``yes``/``on``/
  ``off`` permanecen como **string**; ``true``/``false`` SÍ se
  interpretan como bool.
- Sanidad de ``safe_sheet_title`` y ``file_hash_short``.

NO se verifica contenido semántico de las plantillas (82 preguntas, 14
hitos, fórmulas, formato condicional). Esa validación se hizo
visualmente en el paso 6; replicarla aquí daría tests frágiles ante
cualquier edición legítima de los YAMLs.
"""

from __future__ import annotations

from pathlib import Path

import pytest
import yaml as _pyyaml
from openpyxl import load_workbook

from scripts.render_plantillas import (
    PLANTILLAS_DIR,
    _StrictBoolLoader,
    file_hash_short,
    load_yaml,
    render_cuestionario,
    render_ficha,
    safe_sheet_title,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

CUESTIONARIO_YAML = PLANTILLAS_DIR / "cuestionario_viabilidad.yaml"
INFORME_YAML = PLANTILLAS_DIR / "informe_viabilidad.yaml"


def _read_meta_sheet(xlsx_path: Path) -> dict[str, str]:
    """Lee la hoja oculta ``_meta`` y devuelve sus filas como dict clave→valor."""
    wb = load_workbook(xlsx_path)
    assert "_meta" in wb.sheetnames, "Hoja _meta ausente"
    ws = wb["_meta"]
    pairs: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        key = row[0]
        val = row[1] if len(row) > 1 else None
        if key is not None:
            pairs[str(key)] = "" if val is None else str(val)
    return pairs


# ---------------------------------------------------------------------------
# 1. Smoke — render_cuestionario
# ---------------------------------------------------------------------------

def test_render_cuestionario_genera_xlsx_abrible(tmp_path):
    out = tmp_path / "cuestionario.xlsx"
    render_cuestionario(CUESTIONARIO_YAML, out)

    assert out.is_file()
    wb = load_workbook(out)
    # Hoja _meta presente y oculta
    assert "_meta" in wb.sheetnames
    assert wb["_meta"].sheet_state == "hidden"
    # Hojas de secciones: al menos 1 visible además de _meta
    visible = [s for s in wb.sheetnames if s != "_meta"]
    assert len(visible) >= 1


def test_render_cuestionario_tiene_secciones_del_yaml(tmp_path):
    """Las secciones declaradas en el YAML aparecen como hojas (truncadas
    a 31 chars con safe_sheet_title)."""
    data = load_yaml(CUESTIONARIO_YAML)
    secciones_yaml = data.get("secciones", [])
    assert secciones_yaml, "El YAML del cuestionario no tiene secciones"
    titulos_esperados = {
        safe_sheet_title(s["titulo"]) for s in secciones_yaml
    }

    out = tmp_path / "cuestionario.xlsx"
    render_cuestionario(CUESTIONARIO_YAML, out)
    wb = load_workbook(out)
    hojas_visibles = {s for s in wb.sheetnames if s != "_meta"}

    # Cada título esperado debe estar entre las hojas visibles
    assert titulos_esperados.issubset(hojas_visibles), \
        f"Faltan hojas: {titulos_esperados - hojas_visibles}"


# ---------------------------------------------------------------------------
# 2. Smoke — render_ficha
# ---------------------------------------------------------------------------

def test_render_ficha_genera_xlsx_abrible(tmp_path):
    out = tmp_path / "ficha.xlsx"
    render_ficha(INFORME_YAML, out)

    assert out.is_file()
    wb = load_workbook(out)
    assert "OPERACION" in wb.sheetnames
    assert "_meta" in wb.sheetnames
    assert wb["_meta"].sheet_state == "hidden"


# ---------------------------------------------------------------------------
# 3. Hash del YAML en _meta — detecta plantillas desactualizadas
# ---------------------------------------------------------------------------

def test_meta_de_cuestionario_contiene_hash_correcto_del_yaml(tmp_path):
    out = tmp_path / "cuestionario.xlsx"
    render_cuestionario(CUESTIONARIO_YAML, out)

    meta = _read_meta_sheet(out)
    assert "yaml_sha256_16" in meta
    assert meta["yaml_sha256_16"] == file_hash_short(CUESTIONARIO_YAML)


def test_meta_de_ficha_contiene_hash_correcto_del_yaml(tmp_path):
    out = tmp_path / "ficha.xlsx"
    render_ficha(INFORME_YAML, out)

    meta = _read_meta_sheet(out)
    assert "yaml_sha256_16" in meta
    assert meta["yaml_sha256_16"] == file_hash_short(INFORME_YAML)


# ---------------------------------------------------------------------------
# 4. _StrictBoolLoader — contrato crítico
# ---------------------------------------------------------------------------

def test_strict_bool_loader_no_convierte_si_no_a_bool(tmp_path):
    """``si``/``no``/``yes``/``on``/``off`` deben permanecer como string.

    PyYAML 1.1 los interpreta como bool por defecto — eso rompe el uso
    natural de ``si``/``no`` como valores de respuesta en el cuestionario.
    """
    yaml_text = """
    r_si: si
    r_no: no
    r_yes: yes
    r_on: on
    r_off: off
    """
    src = tmp_path / "inline.yaml"
    src.write_text(yaml_text, encoding="utf-8")
    data = load_yaml(src)

    assert data["r_si"] == "si" and isinstance(data["r_si"], str)
    assert data["r_no"] == "no" and isinstance(data["r_no"], str)
    assert data["r_yes"] == "yes" and isinstance(data["r_yes"], str)
    assert data["r_on"] == "on" and isinstance(data["r_on"], str)
    assert data["r_off"] == "off" and isinstance(data["r_off"], str)


def test_strict_bool_loader_true_false_si_son_bool(tmp_path):
    """``true``/``false`` (y variantes de capitalización) SÍ deben ser bool."""
    yaml_text = """
    activo: true
    inactivo: false
    titulo: True
    falso_cap: FALSE
    """
    src = tmp_path / "inline.yaml"
    src.write_text(yaml_text, encoding="utf-8")
    data = load_yaml(src)

    assert data["activo"] is True
    assert data["inactivo"] is False
    assert data["titulo"] is True
    assert data["falso_cap"] is False


def test_strict_bool_loader_vs_safeloader_demuestra_diferencia(tmp_path):
    """Sanity: el SafeLoader vanilla SÍ convertiría ``si``/``no`` a bool —
    confirmación de que ``_StrictBoolLoader`` está activo, no es no-op."""
    yaml_text = "valor: yes\n"
    src = tmp_path / "inline.yaml"
    src.write_text(yaml_text, encoding="utf-8")

    # SafeLoader vanilla: True (comportamiento por defecto YAML 1.1)
    with open(src, encoding="utf-8") as f:
        vanilla = _pyyaml.safe_load(f)
    assert vanilla["valor"] is True

    # _StrictBoolLoader: string "yes"
    strict = load_yaml(src)
    assert strict["valor"] == "yes"


# ---------------------------------------------------------------------------
# 5. file_hash_short
# ---------------------------------------------------------------------------

def test_file_hash_short_devuelve_16_hex(tmp_path):
    f = tmp_path / "x.bin"
    f.write_bytes(b"contenido cualquiera")
    h = file_hash_short(f)
    assert len(h) == 16
    assert all(c in "0123456789abcdef" for c in h)


def test_file_hash_short_estable_para_mismo_contenido(tmp_path):
    a = tmp_path / "a.bin"
    b = tmp_path / "b.bin"
    a.write_bytes(b"mismo contenido")
    b.write_bytes(b"mismo contenido")
    assert file_hash_short(a) == file_hash_short(b)


def test_file_hash_short_cambia_si_cambia_el_contenido(tmp_path):
    a = tmp_path / "a.bin"
    a.write_bytes(b"v1")
    h1 = file_hash_short(a)
    a.write_bytes(b"v2")
    h2 = file_hash_short(a)
    assert h1 != h2


# ---------------------------------------------------------------------------
# 6. safe_sheet_title — límite 31 + caracteres prohibidos
# ---------------------------------------------------------------------------

def test_safe_sheet_title_trunca_a_31_chars():
    largo = "A" * 100
    assert safe_sheet_title(largo) == "A" * 31


def test_safe_sheet_title_elimina_caracteres_prohibidos():
    sucio = "Sección/con\\caracteres?prohibidos*[a]:b"
    limpio = safe_sheet_title(sucio)
    for bad in ("/", "\\", "?", "*", "[", "]", ":"):
        assert bad not in limpio


def test_safe_sheet_title_pasa_titulo_corto_intacto():
    assert safe_sheet_title("Demanda") == "Demanda"
