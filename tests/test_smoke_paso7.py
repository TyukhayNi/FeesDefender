"""Smoke programático del refactor intake v2 — pasos 7a + 7b.

Sustituye el smoke manual que el usuario no pudo hacer en la sesión
2026-05-11. Cubre el comportamiento end-to-end de las funciones del core
que se invocan tras los clicks en la UI:

- `ensure_case` con `tipo_caso` aplicable e inaplicable.
- Pre-relleno de REF y FECHA en la ficha de operación.
- Idempotencia estricta (segunda llamada no destruye trabajo del abogado).
- `save_file_crm_branch` con rama válida y con path traversal.
- `list_crm_branch_files` sobre rama vacía y rama con contenido.
- `intake_log.append_event` con `upload_manual` apuntando a `05_CRM/...`.

NO cubre:

- Render del sidebar Streamlit (requiere navegador + interacción humana).
- Selector encadenado de CRM_TREE en la UI (requiere Streamlit corriendo).

Esos puntos quedan pendientes de smoke manual cuando el usuario tenga
tiempo de abrir la app. Hasta entonces, este test verifica que el core
hace lo correcto cuando la UI lo invoca con los parámetros documentados.
"""

from __future__ import annotations

import importlib
import json
from datetime import date
from pathlib import Path

import pytest


# ---------------------------------------------------------------------------
# Fixtures auxiliares
# ---------------------------------------------------------------------------

@pytest.fixture
def reloaded_modules(tmp_casos_root):
    """Recarga case_manager / intake_manual / intake_log tras el reload de cfg.

    El fixture `tmp_casos_root` ya recarga `core.config`. Los módulos que
    capturan `settings.casos_root` o `caso_path` en import-time deben
    recargarse también — si no, conservan la referencia al casos_root del
    repo y los tests escriben fuera del tmp.
    """
    from core import config as cfg  # ya recargado por tmp_casos_root
    from core import case_manager, intake_log, intake_manual

    importlib.reload(case_manager)
    importlib.reload(intake_log)
    importlib.reload(intake_manual)

    return {
        "cfg": cfg,
        "case_manager": case_manager,
        "intake_log": intake_log,
        "intake_manual": intake_manual,
    }


# ---------------------------------------------------------------------------
# Paso 7a — ensure_case v2
# ---------------------------------------------------------------------------

def _read_caso_meta(caso_md_path: Path) -> dict:
    import yaml
    text = caso_md_path.read_text(encoding="utf-8")
    assert text.startswith("---"), f"_caso.md sin frontmatter: {caso_md_path}"
    _, fm_raw, _ = text.split("---", 2)
    fm = yaml.safe_load(fm_raw) or {}
    return fm.get("meta") or {}


def test_ensure_case_crm_base_lazy(reloaded_modules, tmp_casos_root):
    """D7 (reorg 2026-06-10) — andamiaje lazy: ensure_case crea solo la base
    ``05_CRM/``, NO el árbol profundo ni los buckets (se materializan al
    escribir). Deroga el antiguo D1 eager.
    """
    cm = reloaded_modules["case_manager"]
    cfg = reloaded_modules["cfg"]

    case_dir = cm.ensure_case("EV-2026-CRM-TREE")
    crm_root = case_dir / "00_Input" / cfg.CRM_SUBDIR

    # La base existe…
    assert crm_root.is_dir()
    # …pero NINGÚN bucket ni rama profunda se pre-crea (no carpetas vacías).
    assert list(crm_root.iterdir()) == [], (
        f"05_CRM no debería tener subcarpetas tras ensure_case: "
        f"{[p.name for p in crm_root.iterdir()]}"
    )
    # Ni las ramas profundas heredadas ni los buckets nuevos.
    for rel in (
        "Civil/1ª Instancia/Declarativo/Demanda",
        "Civil/Preliminares/Demanda",
        "General",
        "01_Demanda",
        "05_Diligencias_Preliminares",
        "99_Otros",
    ):
        assert not (crm_root / rel).exists(), f"No debería existir aún: {rel}"


def test_ensure_case_copia_informe_siempre(reloaded_modules, tmp_casos_root):
    """El informe de viabilidad se copia sin importar el tipo_caso."""
    cm = reloaded_modules["case_manager"]
    case_dir = cm.ensure_case("EV-2026-FICHA-1")
    # case_id legacy → fallback "_informe_viabilidad.xlsx"
    assert (case_dir / "02_Analisis" / "_informe_viabilidad.xlsx").is_file()


def test_ensure_case_copia_cuestionario_solo_si_aplicable(reloaded_modules):
    """El cuestionario solo se copia si tipo_caso ∈ INFORME_VIABILIDAD_TIPOS."""
    cm = reloaded_modules["case_manager"]
    informe_legacy = "_informe_viabilidad.xlsx"   # case_ids legacy → fallback

    # Tipo aplicable
    cd_neg = cm.ensure_case("EV-2026-NEG", tipo_caso="NEGATIVA_OFERTA")
    assert (cd_neg / "02_Analisis" / "_cuestionario_viabilidad.xlsx").is_file()
    assert (cd_neg / "02_Analisis" / informe_legacy).is_file()

    # Tipo NO aplicable
    cd_bd = cm.ensure_case("EV-2026-BAD", tipo_caso="BAD_DEBT")
    assert not (cd_bd / "02_Analisis" / "_cuestionario_viabilidad.xlsx").exists()
    assert (cd_bd / "02_Analisis" / informe_legacy).is_file()

    # Sin tipo_caso
    cd_none = cm.ensure_case("EV-2026-NONE")
    assert not (cd_none / "02_Analisis" / "_cuestionario_viabilidad.xlsx").exists()
    assert (cd_none / "02_Analisis" / informe_legacy).is_file()


def test_ensure_case_prerellena_ref_y_fecha(reloaded_modules):
    """REF y FECHA pre-rellenadas cuando los 3 componentes están presentes."""
    cm = reloaded_modules["case_manager"]
    from openpyxl import load_workbook

    case_id = "BaRS1 - Roger Lluria 38 (W-030LFT) - Negativa Oferta"
    case_dir = cm.ensure_case(
        case_id,
        tipo_caso="NEGATIVA_OFERTA",
        direccion="Roger Lluria 38",
        id_go="W-030LFT",
    )
    # case_id sigue formato CRM nuevo → nombre del informe con solo el ID GO
    informe = case_dir / "02_Analisis" / "Informe viabilidad - W-030LFT.xlsx"
    assert informe.is_file(), f"Informe no encontrado en {informe}"
    wb = load_workbook(informe)
    ws = wb["OPERACION"]

    # Localizar fila por etiqueta en columna B (mismo patrón que el código real)
    ref_row, fecha_row = None, None
    for r in range(1, 200):
        v = ws.cell(row=r, column=2).value
        if v == "REF":
            ref_row = r
        elif v == "FECHA":
            fecha_row = r
    assert ref_row is not None, "Etiqueta REF no encontrada en la ficha"
    assert fecha_row is not None, "Etiqueta FECHA no encontrada en la ficha"

    assert ws.cell(row=ref_row, column=3).value == "BaRS1 - Roger Lluria 38 (W-030LFT)"
    # openpyxl persiste date como datetime al releer del XLSX — comparar el .date()
    fecha_val = ws.cell(row=fecha_row, column=3).value
    if hasattr(fecha_val, "date"):
        fecha_val = fecha_val.date()
    assert fecha_val == date.today()


def test_ensure_case_ref_vacio_si_falta_id_go(reloaded_modules):
    """D-7a-2: REF queda vacío si falta algún componente, sin placeholder."""
    cm = reloaded_modules["case_manager"]
    from openpyxl import load_workbook

    case_id = "BaRS1 - Roger Lluria 38 (W-030LFT) - Negativa Oferta"
    case_dir = cm.ensure_case(
        case_id,
        tipo_caso="NEGATIVA_OFERTA",
        direccion="Roger Lluria 38",
        id_go=None,                    # ← faltante
    )
    # El nombre sale del ID GO que arrastra el case_id, aunque el kwarg falte
    informe = case_dir / "02_Analisis" / "Informe viabilidad - W-030LFT.xlsx"
    wb = load_workbook(informe)
    ws = wb["OPERACION"]
    for r in range(1, 200):
        if ws.cell(row=r, column=2).value == "REF":
            assert ws.cell(row=r, column=3).value in (None, "")
            return
    pytest.fail("Etiqueta REF no encontrada")


def test_ensure_case_persiste_tipo_caso_en_caso_md(reloaded_modules):
    cm = reloaded_modules["case_manager"]
    case_dir = cm.ensure_case(
        "EV-2026-PERSIST",
        tipo_caso="VUELTA",
        direccion="Calle Test 1",
        id_go="W-TEST01",
    )
    meta = _read_caso_meta(case_dir / "00_Input" / "_caso.md")
    assert meta.get("tipo_caso") == "VUELTA"
    assert meta.get("direccion") == "Calle Test 1"
    assert meta.get("id_go") == "W-TEST01"


def test_ensure_case_idempotente_preserva_informe_editado(reloaded_modules):
    """D-7a-4: segunda llamada no sobrescribe informe ni cuestionario."""
    cm = reloaded_modules["case_manager"]
    from openpyxl import load_workbook

    case_dir = cm.ensure_case("EV-2026-IDEM", tipo_caso="NEGATIVA_OFERTA")
    # case_id legacy → fallback "_informe_viabilidad.xlsx"
    informe = case_dir / "02_Analisis" / "_informe_viabilidad.xlsx"

    # El abogado "edita" el informe. Uso una celda fuera del layout principal
    # para no chocar con los rangos merged que crea el render (cabeceras de
    # bloque mergean B-E, así que las celdas E de filas de cabecera están
    # bloqueadas para escritura directa).
    wb = load_workbook(informe)
    ws = wb["OPERACION"]
    ws.cell(row=50, column=7, value="EDIT DEL ABOGADO")
    wb.save(informe)

    # Segunda llamada de ensure_case
    cm.ensure_case("EV-2026-IDEM", tipo_caso="NEGATIVA_OFERTA")

    wb2 = load_workbook(informe)
    assert wb2["OPERACION"].cell(row=50, column=7).value == "EDIT DEL ABOGADO"


def test_ensure_case_no_duplica_informe_con_nombre_legacy(reloaded_modules):
    """Tras acortar el nombre (2026-07-28), un caso ya abierto lleva el informe
    con el case_id completo. ``ensure_case`` debe reconocerlo y NO dejar una
    segunda plantilla en blanco al lado del informe trabajado por el abogado.
    """
    cm = reloaded_modules["case_manager"]
    from openpyxl import load_workbook

    case_id = "BaRS1 - Roger Lluria 38 (W-030LFT) - Negativa Oferta"
    case_dir = cm.ensure_case(case_id, tipo_caso="NEGATIVA_OFERTA")

    # Simular el estado previo: renombrar al nombre largo de antes del cambio.
    analisis = case_dir / "02_Analisis"
    nuevo = analisis / "Informe viabilidad - W-030LFT.xlsx"
    legacy = analisis / f"Informe viabilidad - {case_id}.xlsx"
    nuevo.rename(legacy)

    wb = load_workbook(legacy)
    wb["OPERACION"].cell(row=50, column=7, value="EDIT DEL ABOGADO")
    wb.save(legacy)

    cm.ensure_case(case_id, tipo_caso="NEGATIVA_OFERTA")

    informes = sorted(p.name for p in analisis.glob("*.xlsx")
                      if "cuestionario" not in p.name.lower())
    assert informes == [legacy.name], f"Informe duplicado: {informes}"
    assert load_workbook(legacy)["OPERACION"].cell(row=50, column=7).value == \
        "EDIT DEL ABOGADO"


def test_ensure_case_actualiza_tipo_caso_si_kwarg_difiere(reloaded_modules):
    """D-7a-4: cambiar tipo_caso vía kwarg actualiza el frontmatter."""
    cm = reloaded_modules["case_manager"]

    case_dir = cm.ensure_case("EV-2026-RECLASS", tipo_caso="BAD_DEBT")
    meta1 = _read_caso_meta(case_dir / "00_Input" / "_caso.md")
    assert meta1["tipo_caso"] == "BAD_DEBT"
    assert not (case_dir / "02_Analisis" / "_cuestionario_viabilidad.xlsx").exists()

    cm.ensure_case("EV-2026-RECLASS", tipo_caso="VUELTA")
    meta2 = _read_caso_meta(case_dir / "00_Input" / "_caso.md")
    assert meta2["tipo_caso"] == "VUELTA"
    # El cuestionario aparece porque el destino antes no existía
    assert (case_dir / "02_Analisis" / "_cuestionario_viabilidad.xlsx").is_file()


# ---------------------------------------------------------------------------
# Paso 7b — save_file_crm_branch + intake_log
# ---------------------------------------------------------------------------

def test_save_file_crm_branch_escribe_en_rama_canonica(reloaded_modules):
    cm = reloaded_modules["case_manager"]
    im = reloaded_modules["intake_manual"]

    cm.ensure_case("EV-2026-CRM-UP")
    dest = im.save_file_crm_branch(
        "EV-2026-CRM-UP",
        "Civil/1ª Instancia/Declarativo/Demanda",
        "demanda.pdf",
        b"%PDF-1.4 demo",
    )
    assert dest.is_file()
    assert dest.read_bytes() == b"%PDF-1.4 demo"
    assert dest.name == "demanda.pdf"
    assert "05_CRM" in str(dest)
    assert "Declarativo" in str(dest)


def test_save_file_crm_branch_general_es_destino_valido(reloaded_modules):
    """D-7b-8: subir directamente a 'General' debe funcionar."""
    cm = reloaded_modules["case_manager"]
    im = reloaded_modules["intake_manual"]

    cm.ensure_case("EV-2026-GENERAL")
    dest = im.save_file_crm_branch(
        "EV-2026-GENERAL", "General", "apersonacion.pdf", b"doc",
    )
    assert dest.is_file()
    assert dest.parent.name == "General"


def test_save_file_crm_branch_sobrescribe(reloaded_modules):
    """D-7b-3: filename duplicado se sobrescribe silenciosamente."""
    cm = reloaded_modules["case_manager"]
    im = reloaded_modules["intake_manual"]

    cm.ensure_case("EV-2026-OVR")
    im.save_file_crm_branch("EV-2026-OVR", "General", "x.pdf", b"v1")
    dest = im.save_file_crm_branch("EV-2026-OVR", "General", "x.pdf", b"v2")
    assert dest.read_bytes() == b"v2"


def test_save_file_crm_branch_rechaza_path_traversal(reloaded_modules):
    cm = reloaded_modules["case_manager"]
    im = reloaded_modules["intake_manual"]

    cm.ensure_case("EV-2026-TRAVERSAL")
    with pytest.raises(ValueError):
        im.save_file_crm_branch(
            "EV-2026-TRAVERSAL", "Civil/../../etc", "x.pdf", b"data",
        )


def test_save_file_crm_branch_rechaza_filename_con_separadores(reloaded_modules):
    cm = reloaded_modules["case_manager"]
    im = reloaded_modules["intake_manual"]

    cm.ensure_case("EV-2026-SEPS")
    with pytest.raises(ValueError):
        im.save_file_crm_branch(
            "EV-2026-SEPS", "General", "subdir/x.pdf", b"data",
        )


def test_save_file_crm_branch_caso_inexistente(reloaded_modules):
    im = reloaded_modules["intake_manual"]
    with pytest.raises(FileNotFoundError):
        im.save_file_crm_branch("NO-EXISTE", "General", "x.pdf", b"data")


def test_list_crm_branch_files(reloaded_modules):
    cm = reloaded_modules["case_manager"]
    im = reloaded_modules["intake_manual"]

    cm.ensure_case("EV-2026-LISTC")
    assert im.list_crm_branch_files("EV-2026-LISTC", "General") == []

    im.save_file_crm_branch("EV-2026-LISTC", "General", "a.pdf", b"a")
    im.save_file_crm_branch("EV-2026-LISTC", "General", "b.pdf", b"b")
    files = im.list_crm_branch_files("EV-2026-LISTC", "General")
    assert {f.name for f in files} == {"a.pdf", "b.pdf"}


def test_intake_log_upload_manual_a_05_crm(reloaded_modules, tmp_casos_root):
    """D-7b-5: upload al árbol CRM emite evento upload_manual con details."""
    cm = reloaded_modules["case_manager"]
    im = reloaded_modules["intake_manual"]
    log = reloaded_modules["intake_log"]

    case_id = "EV-2026-LOG"
    cm.ensure_case(case_id)
    log.set_actor("Nikolai Tyukhay")

    branch = "Civil/1ª Instancia/Declarativo/Demanda"
    raw = b"%PDF-1.4 logged"
    im.save_file_crm_branch(case_id, branch, "demanda.pdf", raw)
    log.append_event(
        case_id,
        "upload_manual",
        details={
            "destination": f"05_CRM/{branch}/demanda.pdf",
            "filename": "demanda.pdf",
            "size_bytes": len(raw),
        },
    )

    log_file = tmp_casos_root / case_id / "00_Input" / "_intake_log.jsonl"
    assert log_file.is_file()
    lines = [ln for ln in log_file.read_text(encoding="utf-8").splitlines() if ln.strip()]
    assert len(lines) == 1
    entry = json.loads(lines[0])
    assert entry["event"] == "upload_manual"
    assert entry["actor"] == "Nikolai Tyukhay"
    assert entry["case_id"] == case_id
    assert entry["details"]["destination"].startswith("05_CRM/")
    assert entry["details"]["filename"] == "demanda.pdf"
    assert entry["details"]["size_bytes"] == len(raw)


# ---------------------------------------------------------------------------
# Constante ACTORES_DESPACHO — sanity check
# ---------------------------------------------------------------------------

def test_actores_despacho_lista_minima():
    """Smoke de la constante usada por el sidebar (M10)."""
    from core.config import ACTORES_DESPACHO

    assert isinstance(ACTORES_DESPACHO, tuple)
    assert len(ACTORES_DESPACHO) >= 4
    assert "Nikolai Tyukhay" in ACTORES_DESPACHO
    # Set cerrado, sin duplicados
    assert len(set(ACTORES_DESPACHO)) == len(ACTORES_DESPACHO)
