from pathlib import Path

from core.extractor import _extract_one


def test_extract_rtf(tmp_path: Path):
    p = tmp_path / "burofax.rtf"
    p.write_text(r"{\rtf1\ansi\deff0 Hola \b mundo\b0 burofax\par}", encoding="ascii")
    texto, metodo = _extract_one(p)
    assert metodo == "rtf"
    assert "Hola" in texto and "mundo" in texto and "burofax" in texto


def test_extract_ics(tmp_path: Path):
    p = tmp_path / "invite.ics"
    p.write_text(
        "BEGIN:VCALENDAR\r\n"
        "BEGIN:VEVENT\r\n"
        "SUMMARY:Reunión Tibidabo\r\n"
        "DTSTART:20260604T100000Z\r\n"
        "LOCATION:Barcelona\r\n"
        "END:VEVENT\r\n"
        "END:VCALENDAR\r\n",
        encoding="utf-8",
    )
    texto, metodo = _extract_one(p)
    assert metodo == "ics"
    assert "Reunión Tibidabo" in texto
    assert "DTSTART" in texto and "20260604T100000Z" in texto


def test_extract_xlsm(tmp_path: Path):
    from openpyxl import Workbook

    p = tmp_path / "datos.xlsm"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "hola"
    ws["B1"] = "mundo"
    wb.save(p)  # openpyxl escribe estructura xlsx; la extensión no afecta a la lectura
    texto, metodo = _extract_one(p)
    assert metodo == "pandas"
    assert "hola" in texto and "mundo" in texto
